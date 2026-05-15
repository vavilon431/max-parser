"""
MAX Parser Dashboard — Flask веб-інтерфейс для перегляду matches.db
"""

from flask import Flask, render_template_string, request, jsonify, Response, session, redirect, url_for
from werkzeug.security import check_password_hash
import os
import secrets
import sqlite3
import re
import random
import time
import threading
import io
from pathlib import Path
from datetime import datetime, timedelta

try:
    from nltk.corpus import stopwords as nltk_stopwords
    import nltk
    try:
        _RU_STOPS = set(nltk_stopwords.words("russian"))
    except LookupError:
        nltk.download("stopwords", quiet=True)
        _RU_STOPS = set(nltk_stopwords.words("russian"))
except Exception:
    _RU_STOPS = set()

import nlp as nlp_mod
import uuid

DB_FILE                   = Path(__file__).parent / "matches.db"
CUSTOM_STOPS_FILE         = Path(__file__).parent / "custom_stop_words.txt"
ALERT_CHANNELS_FILE       = Path(__file__).parent / "channels" / "alert_channels.txt"
TOP_WORDS_CACHE_TTL       = 3600   # 1 год — прорахунок на 20k постів коштує дорого
BASELINE_REBUILD_INTERVAL = 24 * 3600
# NER pipeline ~20-40ms/пост (без syntax) → перший прорахунок ~7-14 хв на 20k постах
# для FALLBACK-шляху. Основний шлях — інкрементальний кеш (`message_lemmas`),
# який наповнює `lemma_cache_scheduler` у фоні; запит "топ за 24h" з кеша = мс.
MAX_ROWS_SCAN             = 20_000

# Інкрементальний кеш лем
LEMMA_CACHE_BATCH_SIZE    = 200    # постів за один прохід обробки
LEMMA_CACHE_INTERVAL      = 30     # секунд між проходами (idle)
LEMMA_CACHE_BUSY_INTERVAL = 1      # секунд між проходами поки є pending
LEMMA_CACHE_MIN_COVERAGE  = 0.80   # % постів періоду які мусять бути в кеші,
                                    # щоб не падати в on-the-fly fallback
# Multi-core обробка кеша через ProcessPoolExecutor.
# Кожен worker тримає Natasha (~600 МБ RAM), тож обережно: env override.
# 1 = sequential (за замовчуванням, безпечно для VPS з 800 МБ ліміту).
try:
    LEMMA_CACHE_WORKERS = max(1, int(os.environ.get("MAX_PARSER_NLP_WORKERS", "1")))
except ValueError:
    LEMMA_CACHE_WORKERS = 1

# AI-аналітика через Claude API
ANTHROPIC_KEY_FILE        = Path(__file__).parent / ".anthropic_key"
# Опційний base_url для проксі (Cloudflare AI Gateway / LiteLLM) — потрібен через
# гео-блок Anthropic у RU. Формат: https://gateway.ai.cloudflare.com/v1/<acct>/<gw>/anthropic
ANTHROPIC_GATEWAY_FILE    = Path(__file__).parent / ".anthropic_gateway"
# Опційний токен для Authenticated Gateway (cf-aig-authorization)
ANTHROPIC_GATEWAY_TOKEN_FILE = Path(__file__).parent / ".anthropic_gateway_token"
ANTHROPIC_MODEL           = "claude-sonnet-4-6"
SUMMARY_PROMPT_FILE       = Path(__file__).parent / "summary.txt"
ANALYTICS_MAX_INPUT_CHARS = 600_000   # ~150k токенів — безпечний поріг для context 200k
ANALYTICS_POST_TRIM_CHARS = 1_500     # обрізання дуже довгих постів
ANALYTICS_TASK_TTL        = 1800      # 30 хв
ANALYTICS_CACHE_TTL       = 900       # 15 хв на (q, channel, since_ts, until_ts)

_top_words_cache: dict[tuple[str, str], tuple[float, dict]] = {}
_top_words_inflight: set[tuple[str, str]] = set()  # (period, channel) пари в обчисленні

app = Flask(__name__)


# ── Session-based authentication ─────────────────────────────────────────────
# Файл `.dashboard_auth` — рядки `username:hash` (хеш робить manage_auth.py через
# werkzeug.security.generate_password_hash). Якщо файл відсутній/порожній — auth
# ВИМКНЕНИЙ (для локальної розробки). Сесія тримається 7 днів через signed cookie.
# Secret key зберігається в `.dashboard_secret` (генерується при першому старті).

DASHBOARD_AUTH_FILE   = Path(__file__).parent / ".dashboard_auth"
DASHBOARD_SECRET_FILE = Path(__file__).parent / ".dashboard_secret"
_AUTH_CACHE_TTL       = 60
_auth_cache: tuple[float, dict[str, str]] | None = None
_auth_lock            = threading.Lock()


def _load_or_create_secret() -> bytes:
    if DASHBOARD_SECRET_FILE.exists():
        data = DASHBOARD_SECRET_FILE.read_bytes()
        if len(data) >= 16:
            return data
    key = secrets.token_bytes(32)
    DASHBOARD_SECRET_FILE.write_bytes(key)
    try:
        DASHBOARD_SECRET_FILE.chmod(0o600)
    except OSError:
        pass
    return key


app.secret_key = _load_or_create_secret()
app.permanent_session_lifetime = timedelta(days=7)


def _load_auth_users() -> dict[str, str]:
    global _auth_cache
    now = time.time()
    with _auth_lock:
        if _auth_cache and now - _auth_cache[0] < _AUTH_CACHE_TTL:
            return _auth_cache[1]
    users: dict[str, str] = {}
    if DASHBOARD_AUTH_FILE.exists():
        for line in DASHBOARD_AUTH_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or ":" not in line:
                continue
            user, pwhash = line.split(":", 1)
            user = user.strip()
            if user:
                users[user] = pwhash.strip()
    with _auth_lock:
        _auth_cache = (now, users)
    return users


_PUBLIC_PATHS = {"/login", "/logout"}


@app.before_request
def _require_session_auth():
    users = _load_auth_users()
    if not users:                                  # auth вимкнений
        return None
    if request.path in _PUBLIC_PATHS:
        return None
    if session.get("user") in users:
        return None
    # API-запити отримують 401 JSON, HTML — redirect на /login
    if request.path.startswith("/api/"):
        return jsonify({"error": "auth_required"}), 401
    return redirect(url_for("login_page", next=request.full_path))


_LOGIN_TEMPLATE = """<!doctype html>
<html lang="uk">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Вхід — MAX Radar</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; }
    body {
      margin: 0; min-height: 100vh; font-family: 'Inter', system-ui, sans-serif;
      background: #080b14; color: #c9d1d9;
      display: flex; align-items: center; justify-content: center;
    }
    body::before {
      content: ''; position: fixed; inset: 0; z-index: -1;
      background: radial-gradient(ellipse 60% 40% at 50% 30%, rgba(108,99,255,0.18), transparent),
                  radial-gradient(ellipse 60% 40% at 80% 80%, rgba(56,189,248,0.07), transparent);
    }
    .login-card {
      width: 100%; max-width: 380px; padding: 2rem;
      background: rgba(255,255,255,0.03);
      border: 1px solid rgba(255,255,255,0.08);
      border-radius: 16px;
      box-shadow: 0 10px 40px rgba(0,0,0,0.4);
    }
    .login-title {
      font-size: 0.7rem; font-weight: 700; color: #8b83ff;
      text-transform: uppercase; letter-spacing: 2px; text-align: center; margin-bottom: 0.3rem;
    }
    .login-subtitle {
      font-size: 1.15rem; font-weight: 700; color: #c4bfff;
      text-align: center; margin-bottom: 1.5rem;
    }
    .login-error {
      background: rgba(224,82,82,0.1); border: 1px solid rgba(224,82,82,0.3);
      color: #ff9999; font-size: 0.85rem; padding: 0.6rem 0.85rem;
      border-radius: 8px; margin-bottom: 1rem; text-align: center;
    }
    label { display: block; font-size: 0.7rem; color: #8b83a8; text-transform: uppercase;
      letter-spacing: 1px; margin-bottom: 0.35rem; }
    input[type="text"], input[type="password"] {
      width: 100%; padding: 0.7rem 0.85rem; margin-bottom: 1rem;
      background: rgba(0,0,0,0.3); border: 1px solid rgba(255,255,255,0.1);
      border-radius: 8px; color: #e2e0ff; font-size: 0.95rem;
      transition: border-color 0.15s;
    }
    input:focus { outline: none; border-color: rgba(108,99,255,0.5); }
    button {
      width: 100%; padding: 0.75rem; margin-top: 0.5rem;
      background: linear-gradient(135deg, #6c63ff, #8b83ff);
      color: #fff; border: none; border-radius: 8px;
      font-size: 0.95rem; font-weight: 600; cursor: pointer;
      transition: transform 0.15s, box-shadow 0.15s;
    }
    button:hover { transform: translateY(-1px); box-shadow: 0 6px 18px rgba(108,99,255,0.35); }
    button:active { transform: translateY(0); }
    .login-foot {
      text-align: center; font-size: 0.7rem; color: #4a5568; margin-top: 1.25rem;
    }
  </style>
</head>
<body>
  <form class="login-card" method="post" action="{{ url_for('login_page') }}" autocomplete="off">
    <div class="login-title">MAX Radar</div>
    <div class="login-subtitle">Вхід у дашборд</div>
    {% if error %}<div class="login-error">{{ error }}</div>{% endif %}
    <input type="hidden" name="next" value="{{ next_url }}">
    <label for="u">Логін</label>
    <input id="u" name="username" type="text" required autofocus value="{{ username|e }}">
    <label for="p">Пароль</label>
    <input id="p" name="password" type="password" required>
    <button type="submit">Увійти</button>
    <div class="login-foot">Моніторинг медіа-простору</div>
  </form>
</body>
</html>"""


@app.route("/login", methods=["GET", "POST"])
def login_page():
    users = _load_auth_users()
    next_url = request.args.get("next") or request.form.get("next") or "/"
    # захист від open-redirect — приймаємо тільки відносні шляхи
    if not next_url.startswith("/") or next_url.startswith("//"):
        next_url = "/"
    error = None
    username = ""

    if not users:
        # auth вимкнений (ще не створено жодного юзера) — пускаємо одразу
        return redirect(next_url)

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        if username in users and check_password_hash(users[username], password):
            session.permanent = True
            session["user"] = username
            return redirect(next_url)
        error = "Невірний логін або пароль."

    if session.get("user") in users and request.method == "GET":
        return redirect(next_url)

    return render_template_string(_LOGIN_TEMPLATE,
                                  error=error, username=username, next_url=next_url)


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login_page"))


def load_custom_stops() -> set:
    if not CUSTOM_STOPS_FILE.exists():
        return set()
    return {w.strip().lower() for w in CUSTOM_STOPS_FILE.read_text(encoding="utf-8").splitlines() if w.strip()}

def save_custom_stop(word: str):
    with CUSTOM_STOPS_FILE.open("a", encoding="utf-8") as f:
        f.write(word.strip().lower() + "\n")

TEMPLATE = """
<!DOCTYPE html>
<html lang="uk">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>MAX Radar</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js" defer></script>
  <script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js" defer></script>
  <script src="https://cdn.jsdelivr.net/npm/jspdf@2.5.2/dist/jspdf.umd.min.js" defer></script>
  <style>
    *, *::before, *::after { box-sizing: border-box; }
    :root { --page-max: 1190px; --page-pad: 1.25rem; }
    body { background: #080b14; color: #c9d1d9; font-family: 'Inter', sans-serif; min-height: 100vh; }

    /* ── PDF-режим: інверсна тема для друку ───────────────────────────────── */
    body.pdf-mode { background: #ffffff !important; color: #000 !important; }
    body.pdf-mode::before { display: none !important; }
    body.pdf-mode, body.pdf-mode * {
      color: #000 !important;
      border-color: #cccccc !important;
    }
    body.pdf-mode #report-root,
    body.pdf-mode #pdf-report-header,
    body.pdf-mode .search-wrap,
    body.pdf-mode .result-badge,
    body.pdf-mode .timeline-card,
    body.pdf-mode .analytics-result,
    body.pdf-mode .stat-card,
    body.pdf-mode .topic-card,
    body.pdf-mode .sidebar-card,
    body.pdf-mode .post-card {
      background: #ffffff !important;
      background-image: none !important;
    }
    body.pdf-mode .timeline-reach-total {
      background: #f0fafe !important;
      border-color: #abdfee !important;
    }
    body.pdf-mode .word-bar-track,
    body.pdf-mode .channel-bar-track {
      background: #eeeeee !important;
    }
    body.pdf-mode strong { color: #000 !important; }
    body.pdf-mode mark.highlight { background: #fff59d !important; color: #000 !important; }
    body.pdf-mode .analytics-body .ai-list-item {
      background: #f6f5ff !important; border-left-color: #6c63ff !important;
    }
    body.pdf-mode .analytics-body h3 { color: #000 !important; border-bottom-color: #ccc !important; }
    body.pdf-mode .topics-title strong,
    body.pdf-mode .timeline-title strong,
    body.pdf-mode .channel-name { color: #000 !important; }
    /* Стрічка постів у PDF не потрібна — звіт обмежується блоками до Тематичної аналітики включно. */
    body.pdf-mode #main-grid { display: none !important; }
    /* У PDF контейнери розтягуються на всю доступну ширину знімка html2canvas, щоб
       результат заповнював аркуш A4 без широких білих полів зліва-справа. */
    body.pdf-mode { --page-max: 100% !important; }

    /* Фон з градієнтом */
    body::before {
      content: ''; position: fixed; inset: 0; z-index: -1;
      background: radial-gradient(ellipse 80% 50% at 50% -20%, rgba(108,99,255,0.15), transparent),
                  radial-gradient(ellipse 60% 40% at 80% 80%, rgba(56,189,248,0.06), transparent);
    }

    /* Navbar */
    .topbar {
      background: rgba(13,17,28,0.85); backdrop-filter: blur(12px);
      border-bottom: 1px solid rgba(255,255,255,0.06);
      padding: 0.75rem 0;
      position: sticky; top: 0; z-index: 100;
    }
    .topbar-inner {
      max-width: var(--page-max); margin: 0 auto; padding: 0 var(--page-pad);
      display: flex; align-items: center; justify-content: space-between;
    }
    .brand { font-size: 1rem; font-weight: 700; color: #fff; letter-spacing: -0.3px; display: flex; align-items: center; gap: 8px; }
    .brand-dot { width: 8px; height: 8px; border-radius: 50%; background: #6c63ff; box-shadow: 0 0 8px #6c63ff; animation: pulse 2s infinite; }
    @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.4} }
    .topbar-meta { font-size: 0.78rem; color: #4a5568; }

    /* Stat cards */
    .stat-grid {
      display: grid; grid-template-columns: repeat(4,1fr); gap: 1rem;
      max-width: var(--page-max); margin: 1.5rem auto;
    }
    @media(max-width:768px){ .stat-grid { grid-template-columns: repeat(2,1fr); } }
    .stat-card {
      background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.07);
      border-radius: 14px; padding: 1.25rem 1.5rem;
      transition: transform 0.2s, border-color 0.2s;
      text-align: center;
    }
    .stat-card:hover { transform: translateY(-2px); border-color: rgba(108,99,255,0.4); }
    .stat-icon { font-size: 1.4rem; margin-bottom: 0.5rem; }
    .stat-number { font-size: 1.9rem; font-weight: 700; color: #fff; line-height: 1; }
    .stat-label { font-size: 0.72rem; color: #4a5568; text-transform: uppercase; letter-spacing: 1.2px; margin-top: 4px; }
    .stat-delta { font-size: 0.78rem; color: #6b7280; margin-top: 6px; }
    .stat-delta .ch-count { color: #a0aec0; }
    .stat-delta .delta-up   { color: #4caf50; font-weight: 600; margin-left: 6px; }
    .stat-delta .delta-down { color: #e08080; font-weight: 600; margin-left: 6px; }
    .stat-delta .delta-zero { color: #6b7280; margin-left: 6px; }

    /* Tops row (Топ каналів). Один список — повна ширина, два — 50/50 на широких. */
    .tops-grid {
      display: grid; grid-template-columns: repeat(auto-fit, minmax(380px, 1fr)); gap: 1rem;
      max-width: var(--page-max); margin: 0 auto 1rem;
    }

    /* Topics analytics: 2x2 grid категорій (Персони/Локації/Організації/Терміни) */
    .topics-section {
      max-width: var(--page-max); margin: 0 auto 1.5rem;
    }
    .topics-head {
      display: flex; justify-content: space-between; align-items: center;
      margin-bottom: 0.75rem; flex-wrap: wrap; gap: 0.5rem;
    }
    .topics-title { font-size: 0.68rem; font-weight: 600; color: #8b83ff;
      text-transform: uppercase; letter-spacing: 1.5px; }
    .topics-title strong { color: #c4bfff; }
    .topics-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 1rem; }
    @media(max-width:1100px){ .topics-grid { grid-template-columns: repeat(2, 1fr); } }
    @media(max-width:600px){ .topics-grid { grid-template-columns: 1fr; } }
    .topic-card {
      background: rgba(255,255,255,0.025); border: 1px solid rgba(255,255,255,0.06);
      border-radius: 14px; padding: 1.1rem 1.25rem;
      transition: border-color 0.2s;
      display: flex; flex-direction: column;
    }
    .topic-card:hover { border-color: rgba(108,99,255,0.25); }
    .topic-card-head { display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.85rem; }
    .topic-card-title { font-size: 0.78rem; font-weight: 700; color: #c4bfff;
      letter-spacing: 0.5px; display: flex; align-items: center; gap: 6px; }
    .topic-card-title .topic-emoji { font-size: 0.95rem; }
    .topic-card-count { font-size: 0.68rem; color: #4a5568;
      background: rgba(255,255,255,0.04); padding: 2px 8px; border-radius: 999px; }
    .topic-list {
      display: flex; flex-direction: column; gap: 0.35rem;
    }
    .topic-empty { font-size: 0.78rem; color: #3d4a5c; text-align: center;
      padding: 1.5rem 0.5rem; font-style: italic; }
    .topic-card[data-cat="per"]  .word-row-link { color: #ffb86c; }
    .topic-card[data-cat="per"]  .word-row-link:hover { color: #ffd7a8; }
    .topic-card[data-cat="per"]  .word-bar-fill { background: linear-gradient(90deg, #ffb86c, #ffd7a8); }
    .topic-card[data-cat="loc"]  .word-row-link { color: #56cfe1; }
    .topic-card[data-cat="loc"]  .word-row-link:hover { color: #90e0ef; }
    .topic-card[data-cat="loc"]  .word-bar-fill { background: linear-gradient(90deg, #56cfe1, #90e0ef); }
    .topic-card[data-cat="org"]  .word-row-link { color: #a09fff; }
    .topic-card[data-cat="org"]  .word-row-link:hover { color: #c4bfff; }
    .topic-card[data-cat="org"]  .word-bar-fill { background: linear-gradient(90deg, #6c63ff, #a09fff); }
    .topic-card[data-cat="term"] .word-row-link { color: #50c878; }
    .topic-card[data-cat="term"] .word-row-link:hover { color: #82e0a8; }
    .topic-card[data-cat="term"] .word-bar-fill { background: linear-gradient(90deg, #50c878, #82e0a8); }

    /* AI-аналітика (контейнер замість стрічки постів) */
    .analytics-result {
      max-width: var(--page-max); margin: 0 auto 1.5rem; padding: 1.25rem;
      background: linear-gradient(135deg, rgba(108,99,255,0.08), rgba(139,131,255,0.04));
      border: 1px solid rgba(108,99,255,0.25); border-radius: 14px;
    }
    .analytics-head {
      display: flex; justify-content: space-between; align-items: center;
      margin-bottom: 0.75rem; gap: 0.75rem;
    }
    .analytics-title { font-size: 1.05rem; font-weight: 700; color: #c4bfff; }
    .analytics-title span { color: #ffd7a8; }
    .analytics-close {
      background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1);
      color: #c4bfff; border-radius: 8px; padding: 4px 10px; cursor: pointer;
      font-size: 0.85rem; transition: all 0.15s;
    }
    .analytics-close:hover { background: rgba(255,255,255,0.1); color: #fff; }
    .analytics-meta {
      font-size: 0.78rem; color: #8b83a8; margin-bottom: 1rem;
      padding-bottom: 0.6rem; border-bottom: 1px solid rgba(255,255,255,0.05);
    }
    .analytics-body { font-size: 0.92rem; line-height: 1.55; color: #e2e0ff; }
    .analytics-body h3 {
      font-size: 1rem; font-weight: 700; color: #ffb86c;
      margin: 1rem 0 0.5rem; padding-bottom: 0.3rem;
      border-bottom: 1px solid rgba(255,184,108,0.2);
    }
    .analytics-body h3:first-child { margin-top: 0; }
    .analytics-body p { margin: 0.5rem 0; }
    .analytics-body strong { color: #c4bfff; font-weight: 600; }
    .analytics-body .ai-list-item {
      margin: 0.4rem 0; padding: 0.5rem 0.75rem;
      background: rgba(255,255,255,0.025); border-left: 3px solid #6c63ff;
      border-radius: 6px;
    }

    /* Main layout */
    .main-grid {
      display: grid; grid-template-columns: 1fr; gap: 1.5rem;
      max-width: var(--page-max); margin: 0 auto; padding: 0 0 2rem;
    }

    /* Search bar */
    .search-wrap {
      background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.07);
      border-radius: 14px; padding: 1rem 1.25rem; margin-bottom: 1rem;
    }
    .form-control, .form-select {
      background: rgba(0,0,0,0.3) !important; border: 1px solid rgba(255,255,255,0.1) !important;
      color: #c9d1d9 !important; border-radius: 8px !important; font-size: 0.875rem;
    }
    .form-control:focus, .form-select:focus {
      border-color: rgba(108,99,255,0.6) !important; box-shadow: 0 0 0 3px rgba(108,99,255,0.15) !important;
    }
    .form-control::placeholder { color: #4a5568; }
    .form-select option { background: #0d111c; }
    .btn-search {
      background: linear-gradient(135deg, #6c63ff, #5a52e0); border: none;
      color: #fff; font-weight: 600; font-size: 0.875rem; border-radius: 8px; padding: 0.5rem 1.25rem;
      transition: opacity 0.2s, transform 0.1s;
    }
    .btn-search:hover { opacity: 0.9; transform: translateY(-1px); color: #fff; }
    .btn-clear { background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1); color: #888; border-radius: 8px; padding: 0.5rem 0.75rem; font-size: 0.875rem; transition: background 0.2s; }
    .btn-clear:hover { background: rgba(255,255,255,0.1); color: #ccc; }

    /* Channel autocomplete */
    .channel-ac { position: relative; flex: 1; min-width: 160px; max-width: 260px; }
    .channel-ac-input { width: 100%; }
    .channel-ac-clear {
      position: absolute; right: 8px; top: 50%; transform: translateY(-50%);
      background: transparent; border: none; color: #4a5568; font-size: 1rem;
      cursor: pointer; padding: 0 4px; line-height: 1; display: none;
    }
    .channel-ac-clear:hover { color: #c4bfff; }
    .channel-ac.has-value .channel-ac-clear { display: block; }
    .channel-ac-list {
      position: absolute; top: calc(100% + 4px); left: 0; right: 0;
      max-height: 320px; overflow-y: auto;
      background: #0d111c; border: 1px solid rgba(108,99,255,0.3);
      border-radius: 8px; box-shadow: 0 8px 24px rgba(0,0,0,0.4);
      z-index: 200; display: none;
    }
    .channel-ac.open .channel-ac-list { display: block; }
    .channel-ac-item {
      padding: 8px 12px; font-size: 0.85rem; color: #c9d1d9;
      cursor: pointer; transition: background 0.1s;
      white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    }
    .channel-ac-item:hover, .channel-ac-item.active {
      background: rgba(108,99,255,0.18); color: #fff;
    }
    .channel-ac-item mark {
      background: rgba(108,99,255,0.35); color: #fff; padding: 0 2px; border-radius: 3px;
    }
    .channel-ac-empty {
      padding: 10px 12px; font-size: 0.8rem; color: #4a5568; font-style: italic;
    }
    .btn-refresh {
      background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1);
      color: #888; font-size: 0.78rem; border-radius: 8px; padding: 5px 12px;
      transition: all 0.2s; text-decoration: none;
    }
    .btn-refresh:hover { background: rgba(108,99,255,0.15); border-color: rgba(108,99,255,0.4); color: #a09fff; }

    /* Preset chips (часові фільтри) */
    .preset-chips { display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 4px; }
    .preset-chip {
      background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.08);
      color: #6b7280; font-size: 0.78rem; font-weight: 600;
      border-radius: 999px; padding: 5px 14px; cursor: pointer;
      transition: all 0.15s; line-height: 1;
    }
    .preset-chip:hover { background: rgba(108,99,255,0.1); color: #a09fff; border-color: rgba(108,99,255,0.3); }
    .preset-chip.active { background: rgba(108,99,255,0.2); border-color: rgba(108,99,255,0.5); color: #c4bfff; }

    /* Custom dates row */
    .custom-dates { display: none; gap: 8px; align-items: center; padding: 6px 0; }
    .custom-dates.show { display: flex; flex-wrap: wrap; }
    .dash-sep { color: #4a5568; font-weight: 600; }

    /* Result badge */
    .result-badge {
      background: linear-gradient(135deg, rgba(108,99,255,0.08), rgba(56,189,248,0.04));
      border: 1px solid rgba(108,99,255,0.2);
      border-radius: 14px; padding: 1rem 1.25rem; margin-bottom: 1rem;
      display: flex; align-items: center; flex-wrap: wrap; gap: 0.75rem 1.5rem;
    }
    .result-badge-main { display: flex; align-items: baseline; gap: 0.5rem; flex: 0 0 auto; }
    .result-badge-icon { font-size: 1.1rem; }
    .result-badge-count { font-size: 1.5rem; font-weight: 700; color: #fff; line-height: 1; font-variant-numeric: tabular-nums; }
    .result-badge-label { font-size: 0.85rem; color: #8b83ff; font-weight: 500; }
    .result-badge-meta { font-size: 0.78rem; color: #6b7280; flex: 1 1 auto; min-width: 0; }
    .result-badge-meta strong { color: #c4bfff; font-weight: 600; }
    .result-badge-sparkline { flex: 0 0 auto; }
    .sparkline { display: block; }
    .sparkline rect { transition: opacity 0.15s; }
    .sparkline rect:hover { opacity: 0.7; }

    /* Timeline (динаміка згадок) */
    .timeline-card {
      background: rgba(255,255,255,0.025); border: 1px solid rgba(255,255,255,0.06);
      border-radius: 14px; padding: 1rem 1.25rem; margin-bottom: 1rem;
    }
    .timeline-head { display: flex; justify-content: space-between; align-items: center; margin-bottom: 0.75rem; flex-wrap: wrap; gap: 0.5rem; }
    .timeline-reach-total { font-size: 0.78rem; font-weight: 500; color: #56cfe1;
      margin-left: 0.5rem; padding: 2px 8px; background: rgba(86,207,225,0.08);
      border: 1px solid rgba(86,207,225,0.25); border-radius: 999px; }
    .timeline-reach-total:empty { display: none; }
    .timeline-title { font-size: 0.85rem; color: #8b83ff; font-weight: 600; }
    .timeline-title strong { color: #c4bfff; }
    .timeline-tabs { display: flex; gap: 4px; }
    .timeline-tab {
      background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.07);
      color: #6b7280; font-size: 0.72rem; font-weight: 600;
      border-radius: 999px; padding: 4px 12px; cursor: pointer;
      transition: all 0.15s; line-height: 1;
    }
    .timeline-tab:hover { color: #a09fff; border-color: rgba(108,99,255,0.3); }
    .timeline-tab.active { background: rgba(108,99,255,0.2); border-color: rgba(108,99,255,0.5); color: #c4bfff; }
    .timeline-canvas-wrap { position: relative; height: 220px; }
    .reach-status {
      font-size: 0.72rem; color: #6b7280; padding: 6px 10px; margin-bottom: 8px;
      background: rgba(108,99,255,0.05); border: 1px solid rgba(108,99,255,0.15);
      border-radius: 8px;
    }
    .reach-status.error { color: #e08080; background: rgba(224,82,82,0.07); border-color: rgba(224,82,82,0.25); }
    .reach-status.done  { color: #8b83ff; }

    /* Post cards */
    .post-card {
      background: rgba(255,255,255,0.025); border: 1px solid rgba(255,255,255,0.06);
      border-radius: 12px; padding: 1rem 1.25rem; margin-bottom: 0.65rem;
      transition: border-color 0.2s, background 0.2s; position: relative; overflow: hidden;
    }
    .post-card::before {
      content: ''; position: absolute; left: 0; top: 0; bottom: 0; width: 3px;
      background: transparent; border-radius: 3px 0 0 3px; transition: background 0.2s;
    }
    .post-card:hover { border-color: rgba(108,99,255,0.3); background: rgba(108,99,255,0.04); }
    .post-card:hover::before { background: #6c63ff; }

    .channel-name { color: #8b83ff; font-weight: 600; font-size: 0.82rem; text-decoration: none !important; transition: color 0.15s; }
    .channel-name:hover { color: #a09fff; }
    .subs-badge { background: rgba(255,255,255,0.05); color: #4a5568; font-size: 0.68rem; padding: 2px 7px; border-radius: 20px; border: 1px solid rgba(255,255,255,0.07); }
    .post-time { font-size: 0.73rem; color: #3d4a5c; font-variant-numeric: tabular-nums; }
    .post-text { font-size: 0.9rem; line-height: 1.6; color: #b0bec5; margin: 0.5rem 0; }
    .post-saved { font-size: 0.7rem; color: #2d3748; }
    .post-link { font-size: 0.72rem; color: #3d4a5c; text-decoration: none; transition: color 0.15s; }
    .post-link:hover { color: #8b83ff; }

    mark.highlight { background: rgba(108,99,255,0.25); color: #c4bfff; border-radius: 3px; padding: 0 3px; font-weight: 600; }

    /* Pagination */
    .pagination { gap: 3px; }
    .page-link { background: rgba(255,255,255,0.03) !important; border: 1px solid rgba(255,255,255,0.08) !important; color: #4a5568 !important; border-radius: 7px !important; font-size: 0.8rem; padding: 5px 11px; transition: all 0.15s; }
    .page-link:hover { background: rgba(108,99,255,0.2) !important; color: #a09fff !important; border-color: rgba(108,99,255,0.3) !important; }
    .page-item.active .page-link { background: #6c63ff !important; border-color: #6c63ff !important; color: #fff !important; }

    /* Sidebar */
    .sidebar-card {
      background: rgba(255,255,255,0.025); border: 1px solid rgba(255,255,255,0.06);
      border-radius: 14px; padding: 1.25rem; overflow-y: auto; max-height: 560px;
    }
    .sidebar-title { font-size: 0.68rem; font-weight: 600; color: #3d4a5c; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 1rem; }
    .channel-row { margin-bottom: 0.85rem; }
    .channel-row-link { color: #8b83ff; font-size: 0.8rem; font-weight: 500; text-decoration: none !important; display: block; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 220px; transition: color 0.15s; }
    .channel-row-link:hover { color: #c4bfff; }
    .channel-row-cnt { font-size: 0.72rem; color: #3d4a5c; white-space: nowrap; }
    .channel-bar-track { background: rgba(255,255,255,0.04); border-radius: 4px; height: 3px; margin-top: 4px; }
    .channel-bar-fill { height: 3px; border-radius: 4px; background: linear-gradient(90deg, #6c63ff, #8b83ff); }

    /* Period tabs */
    .period-tabs { display: flex; gap: 4px; margin-bottom: 1rem; }
    .period-tab {
      flex: 1; text-align: center; font-size: 0.68rem; font-weight: 600; padding: 4px 0;
      border-radius: 6px; border: 1px solid rgba(255,255,255,0.07);
      color: #4a5568; text-decoration: none !important; transition: all 0.15s;
      background: rgba(255,255,255,0.02); cursor: pointer;
    }
    .period-tab:hover { color: #a09fff; border-color: rgba(108,99,255,0.3); }
    .period-tab.active { background: rgba(108,99,255,0.2); border-color: rgba(108,99,255,0.5); color: #c4bfff; }

    /* Word row */
    .word-row { margin-bottom: 0.65rem; }
    .word-row-link { color: #56cfe1; font-size: 0.8rem; font-weight: 500; text-decoration: none !important; display: block; transition: color 0.15s; }
    .word-row-link:hover { color: #90e0ef; }
    .word-row-cnt { font-size: 0.72rem; color: #3d4a5c; white-space: nowrap; }
    .word-bar-track { background: rgba(255,255,255,0.04); border-radius: 4px; height: 3px; margin-top: 4px; }
    .word-bar-fill { height: 3px; border-radius: 4px; background: linear-gradient(90deg, #56cfe1, #90e0ef); }

    /* Empty state */
    .empty-state { background: rgba(255,255,255,0.02); border: 1px dashed rgba(255,255,255,0.08); border-radius: 14px; padding: 3rem; text-align: center; color: #3d4a5c; }

    /* Stop-word button */
    .stop-btn {
      background: none; border: none; color: #2d3748; font-size: 0.7rem; padding: 0 3px;
      cursor: pointer; line-height: 1; transition: color 0.15s; flex-shrink: 0;
    }
    .stop-btn:hover { color: #e05252; }

    /* Modal */
    .stop-modal-overlay {
      display: none; position: fixed; inset: 0; z-index: 500;
      background: rgba(0,0,0,0.6); backdrop-filter: blur(4px);
      align-items: center; justify-content: center;
    }
    .stop-modal-overlay.show { display: flex; }
    .stop-modal-box {
      background: #0d111c; border: 1px solid rgba(255,255,255,0.1);
      border-radius: 16px; padding: 1.75rem 2rem; max-width: 360px; width: 90%;
      box-shadow: 0 20px 60px rgba(0,0,0,0.5);
    }
    .stop-modal-title { font-size: 1rem; font-weight: 700; color: #fff; margin-bottom: 0.5rem; }
    .stop-modal-desc  { font-size: 0.85rem; color: #6b7280; margin-bottom: 1.5rem; }
    .stop-modal-desc strong { color: #c4bfff; }
    .stop-modal-actions { display: flex; gap: 0.75rem; justify-content: flex-end; }
    .stop-modal-no  {
      background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1);
      color: #9ca3af; border-radius: 8px; padding: 0.45rem 1.25rem; font-size: 0.875rem; cursor: pointer;
      transition: background 0.15s;
    }
    .stop-modal-no:hover { background: rgba(255,255,255,0.1); color: #d1d5db; }
    .stop-modal-yes {
      background: linear-gradient(135deg, #e05252, #c03030); border: none;
      color: #fff; border-radius: 8px; padding: 0.45rem 1.25rem; font-size: 0.875rem;
      font-weight: 600; cursor: pointer; transition: opacity 0.15s;
    }
    .stop-modal-yes:hover { opacity: 0.85; }

    ::-webkit-scrollbar { width: 5px; }
    ::-webkit-scrollbar-track { background: transparent; }
    ::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.08); border-radius: 5px; }
  </style>
</head>
<body>

<!-- Topbar -->
<div class="topbar">
  <div class="topbar-inner">
    <div class="brand">
      <div class="brand-dot"></div>
      MAX Radar
    </div>
    <div class="d-flex align-items-center gap-3">
      <span class="topbar-meta">{{ stats.total }} постів · {{ now }}</span>
      <a href="/" class="btn-refresh" onclick="this.textContent='оновлення...'; this.style.pointerEvents='none'">⟳ Оновити</a>
      {% if current_user %}<a href="/logout" class="btn-refresh" title="Вийти з акаунту {{ current_user }}">⎋ {{ current_user }}</a>{% endif %}
    </div>
  </div>
</div>

<div id="report-root">

<!-- Stat cards -->
<div class="stat-grid">
  <div class="stat-card">
    <div class="stat-icon">📄</div>
    <div class="stat-number">{{ stats.total }}</div>
    <div class="stat-label">Всього постів</div>
  </div>
  <div class="stat-card">
    <div class="stat-icon">📡</div>
    <div class="stat-number">{{ stats.channels }}</div>
    <div class="stat-label">Каналів</div>
    <div class="stat-delta">
      <span class="ch-count">{{ stats.active_channels_24h }} активних за 24h</span>
      {% set dch = stats.active_channels_24h - stats.active_channels_24h_prev %}
      {% if dch > 0 %}<span class="delta-up">↑+{{ dch }}</span>
      {% elif dch < 0 %}<span class="delta-down">↓{{ dch }}</span>
      {% else %}<span class="delta-zero">·</span>{% endif %}
    </div>
  </div>
  <div class="stat-card">
    <div class="stat-icon">⚡</div>
    <div class="stat-number">{{ stats.last_hour }}</div>
    <div class="stat-label">За годину</div>
    <div class="stat-delta">
      {% set dh = stats.last_hour - stats.last_hour_prev %}
      {% if dh > 0 %}<span class="delta-up">↑+{{ dh }}</span>
      {% elif dh < 0 %}<span class="delta-down">↓{{ dh }}</span>
      {% else %}<span class="delta-zero">·</span>{% endif %}
    </div>
  </div>
  <div class="stat-card">
    <div class="stat-icon">📅</div>
    <div class="stat-number">{{ stats.last_day }}</div>
    <div class="stat-label">За 24 год</div>
    <div class="stat-delta">
      {% set dd = stats.last_day - stats.last_day_prev %}
      {% if dd > 0 %}<span class="delta-up">↑+{{ dd }}</span>
      {% elif dd < 0 %}<span class="delta-down">↓{{ dd }}</span>
      {% else %}<span class="delta-zero">·</span>{% endif %}
    </div>
  </div>
</div>

<!-- Search -->
<div class="search-wrap" style="max-width: var(--page-max); margin: 0 auto 1rem; padding: 1rem 1.25rem;">
  <form method="get" id="search-form" class="d-flex gap-2 flex-wrap">
    <!-- Єдине джерело правди про period для всієї форми; чіпи лише оновлюють це поле -->
    <input type="hidden" name="period" id="period-hidden" value="{{ period }}">

    <!-- Ряд 1: пресет-чіпи -->
    <div class="preset-chips" style="width:100%">
      {% for p, label in [('1h','Година'),('24h','Доба'),('7d','Тиждень'),('30d','Місяць'),('all','Весь час')] %}
      <button type="button" data-period="{{ p }}" class="preset-chip period-chip {% if period==p %}active{% endif %}">{{ label }}</button>
      {% endfor %}
      <button type="button" id="period-chip-custom" class="preset-chip {% if period=='custom' %}active{% endif %}" onclick="toggleCustomDates()">📅 Інше</button>
    </div>

    <!-- Ряд 2: кастомний діапазон (схований за замовчуванням) -->
    <div id="custom-dates" class="custom-dates {% if period=='custom' %}show{% endif %}" style="width:100%">
      <input type="date" name="from_date" class="form-control" value="{{ from_date }}" style="flex:0 0 auto">
      <span class="dash-sep">—</span>
      <input type="date" name="to_date"   class="form-control" value="{{ to_date }}"   style="flex:0 0 auto">
      <button type="submit" class="btn-search">Застосувати</button>
    </div>

    <!-- Ряд 3: q + channel + sort -->
    <input type="text" name="q" class="form-control" style="flex:1;min-width:180px;"
           placeholder='🔍  Пошук (кілька слів = AND; OR, NOT, "точна фраза", префікс*)'
           title='Приклади: путин лавров — оба слова; путин OR зеленский; "ядерное оружие" — фраза; крым* — префікс; путин NOT медведев'
           value="{{ q }}">
    <div class="channel-ac {% if channel %}has-value{% endif %}" id="channel-ac">
      <input type="text" name="channel" id="channel-ac-input"
             class="form-control channel-ac-input" autocomplete="off"
             placeholder="Всі канали (почніть вводити назву)"
             value="{{ channel }}">
      <button type="button" class="channel-ac-clear" id="channel-ac-clear"
              title="Скинути канал">✕</button>
      <div class="channel-ac-list" id="channel-ac-list"></div>
    </div>
    <select name="sort" class="form-select" style="flex:0 0 auto;width:130px" title="Сортування">
      <option value="new" {% if sort=='new' %}selected{% endif %}>Спочатку нові</option>
      <option value="old" {% if sort=='old' %}selected{% endif %}>Спочатку старі</option>
    </select>
    <button type="submit" class="btn-search">Знайти</button>
    <button type="button" class="btn-search" id="export-xlsx-btn"
            title="Завантажити поточну вибірку в .xlsx (актуальна кількість переглядів збирається з MAX на момент завантаження)"
            aria-label="Завантажити в .xlsx">💾</button>
    {# Тимчасово прихована: тримаємо в DOM (живий JS-обробник `analytics-btn`),
       але візуально не показуємо до подальшого рішення. #}
    <button type="button" class="btn-search" id="analytics-btn" disabled
            style="display:none"
            title="Тимчасово недоступно">🧠 Аналітика</button>
    <button type="button" class="btn-search" onclick="generatePDF(this)"
            title="Сформувати PDF-звіт того, що зараз на дашборді">📄 Звіт</button>
    {# Скидаємо лише q+channel, period (і from/to для custom) лишаємо поточним. #}
    {% if q or channel %}
      {% set _clear_qs = '?period=' ~ period %}
      {% if period == 'custom' %}{% set _clear_qs = _clear_qs ~ '&from_date=' ~ from_date ~ '&to_date=' ~ to_date %}{% endif %}
      <a href="/{{ _clear_qs }}" class="btn-clear" title="Скинути запит і канал (період зберігається)">✕</a>
    {% endif %}
  </form>
</div>

<!-- Result badge -->
<div class="result-badge" style="max-width: var(--page-max); margin: 0 auto 1rem; padding: 1rem 1.25rem;">
  <div class="result-badge-main">
    <span class="result-badge-icon">🔎</span>
    <span class="result-badge-count">{{ "{:,}".format(total_count).replace(",", " ") }}</span>
    <span class="result-badge-label">публікацій</span>
  </div>
  <div class="result-badge-meta">
    {% if q %}для запиту <strong>«{{ q }}»</strong>{% endif %}
    {% if channel %}{% if q %}·{% endif %} канал <strong>{{ channel }}</strong>{% endif %}
    {% if period_label %}{% if q or channel %}·{% endif %} {{ period_label }}{% if period_dates %} ({{ period_dates }}){% endif %}{% endif %}
  </div>
  {% if sparkline_svg %}<div class="result-badge-sparkline">{{ sparkline_svg | safe }}</div>{% endif %}
</div>

<!-- Динаміка згадок (інтерактивний графік) -->
<div class="timeline-card" id="timeline-card" data-q="{{ q }}" data-channel="{{ channel }}"
     style="max-width: var(--page-max); margin: 0 auto 1.5rem; padding: 1rem 1.25rem;">
  <div class="timeline-head">
    <div class="timeline-title">
      Динаміка згадок{% if q %} <strong>«{{ q }}»</strong>{% endif %}{% if channel %} · {{ channel }}{% endif %}
      <span id="timeline-reach-total" class="timeline-reach-total"></span>
    </div>
    <div class="timeline-tabs">
      {# `active`-клас навмисно не виставлений у шаблоні: вибраний період
         відновлюється з localStorage у DOMContentLoaded (default 7). #}
      <button type="button" class="timeline-tab" data-days="7">7 днів</button>
      <button type="button" class="timeline-tab" data-days="30">30 днів</button>
      <button type="button" class="timeline-tab" data-days="90">90 днів</button>
    </div>
  </div>
  <div id="reach-status" class="reach-status" style="display:none"></div>
  <div class="timeline-canvas-wrap"><canvas id="timeline-chart"></canvas></div>
</div>

<!-- Tops row: Топ каналів. Без q — два списки (основний + alert), з q — один. -->
<div class="tops-grid">
  <div class="sidebar-card">
    <div class="sidebar-title">{% if top_channels_alert is not none %}Топ каналів — основний потік{% else %}Топ каналів{% endif %} ({{ top_channels_total }})</div>
    {% set max_cnt = top_channels[0].cnt if top_channels else 1 %}
    {% for ch in top_channels %}
    {% set display_name = ch.channel_title if ch.channel_title and not ch.channel_title.lstrip('-').isdigit() else ch.channel_link.split('/')[-1] %}
    <div class="channel-row">
      <div class="d-flex justify-content-between align-items-center">
        <a href="/?channel={{ ch.channel_title|urlencode }}" class="channel-row-link">{{ display_name }}</a>
        <span class="channel-row-cnt">{{ ch.cnt }}</span>
      </div>
      <div class="channel-bar-track">
        <div class="channel-bar-fill" style="width:{{ (ch.cnt / max_cnt * 100)|int }}%"></div>
      </div>
    </div>
    {% endfor %}
  </div>
  {% if top_channels_alert is not none %}
  <div class="sidebar-card">
    <div class="sidebar-title">Топ каналів — БПЛА / тривоги / радари ({{ top_channels_alert_total }})</div>
    {% if top_channels_alert %}
    {% set max_cnt_alert = top_channels_alert[0].cnt %}
    {% for ch in top_channels_alert %}
    {% set display_name = ch.channel_title if ch.channel_title and not ch.channel_title.lstrip('-').isdigit() else ch.channel_link.split('/')[-1] %}
    <div class="channel-row">
      <div class="d-flex justify-content-between align-items-center">
        <a href="/?channel={{ ch.channel_title|urlencode }}" class="channel-row-link">{{ display_name }}</a>
        <span class="channel-row-cnt">{{ ch.cnt }}</span>
      </div>
      <div class="channel-bar-track">
        <div class="channel-bar-fill" style="width:{{ (ch.cnt / max_cnt_alert * 100)|int }}%"></div>
      </div>
    </div>
    {% endfor %}
    {% else %}
    <div class="topic-empty">Список <code>channels/alert_channels.txt</code> порожній або без активності за період</div>
    {% endif %}
  </div>
  {% endif %}
</div>

<!-- Topics analytics: 2x2 grid (Персони / Локації / Організації / Терміни) -->
<div class="topics-section" data-channel="{{ channel }}">
  <div class="topics-head">
    <div class="topics-title">
      Тематична аналітика{% if words_period_label %} {{ words_period_label }}{% endif %} — <strong>військово-політична</strong>{% if channel %} · канал <strong>{{ channel }}</strong>{% endif %}
    </div>
  </div>
  <div class="topics-grid">
    {% set _topic_meta = [
      ('per',  '👤', 'Персони'),
      ('loc',  '📍', 'Локації'),
      ('org',  '🏛️', 'Організації'),
      ('term', '⚔️', 'Терміни')
    ] %}
    {% for cat, emoji, title in _topic_meta %}
    <div class="topic-card" data-cat="{{ cat }}">
      <div class="topic-card-head">
        <div class="topic-card-title"><span class="topic-emoji">{{ emoji }}</span> {{ title }}</div>
        <div class="topic-card-count" id="topic-count-{{ cat }}">{{ top_words[cat]|length }}</div>
      </div>
      <div class="topic-list" id="topic-list-{{ cat }}">
        {% set rows = top_words[cat] %}
        {% if rows %}
          {% set max_score = rows[0][2] if rows[0][2] > 0 else 1 %}
          {% for word, cnt, score in rows %}
          <div class="word-row">
            <div class="d-flex justify-content-between align-items-center">
              <span class="d-flex align-items-center gap-2" style="min-width:0">
                <a href="/?q={{ word|urlencode }}" class="word-row-link">{{ word }}</a>
                <button class="stop-btn" title="Додати до стоп-слів" onclick='addStopWord(event, {{ word|tojson }})'>✕</button>
              </span>
              <span class="word-row-cnt" title="кількість згадок">{{ cnt }}</span>
            </div>
            <div class="word-bar-track">
              <div class="word-bar-fill" style="width:{{ (score / max_score * 100)|int }}%"></div>
            </div>
          </div>
          {% endfor %}
        {% else %}
          <div class="topic-empty">{% if nlp_ready %}поки немає даних{% else %}NLP-пайплайн ініціалізується…{% endif %}</div>
        {% endif %}
      </div>
    </div>
    {% endfor %}
  </div>
</div>

<!-- AI-аналітика (показується замість стрічки після натискання «🧠 Аналітика») -->
<div id="analytics-result" class="analytics-result" style="display:none">
  <div class="analytics-head">
    <div class="analytics-title">🧠 AI-аналітика — <span id="analytics-q"></span></div>
    <button type="button" class="analytics-close" onclick="closeAnalytics()"
            title="Закрити і повернути стрічку постів">✕</button>
  </div>
  <div id="analytics-meta" class="analytics-meta"></div>
  <div id="analytics-body" class="analytics-body"></div>
</div>

<!-- Main -->
<div class="main-grid" id="main-grid">

  <!-- Feed -->
  <div>
    {% if posts %}
    {% for p in posts %}
    <div class="post-card">
      <div class="d-flex justify-content-between align-items-center mb-1">
        <div class="d-flex align-items-center gap-2">
          <a href="{{ p.channel_link }}" target="_blank" class="channel-name">{{ p.channel_title if p.channel_title and not p.channel_title.lstrip('-').isdigit() else p.channel_link.split('/')[-1] }}</a>
          <span class="subs-badge">{{ "{:,}".format(p.channel_subs) }} підп.</span>
        </div>
        <span class="post-time">{{ p.msg_time }}</span>
      </div>
      <div class="post-text">{{ p.text_hl | safe }}</div>
      <div class="d-flex justify-content-between align-items-center mt-1">
        <span class="post-saved">збережено: {{ p.saved_at }}</span>
        <a href="{{ p.post_link }}" target="_blank" class="post-link">відкрити пост →</a>
      </div>
    </div>
    {% endfor %}

    {% if total_pages > 1 %}
    {% set qs = 'q=' ~ q ~ '&channel=' ~ channel ~ '&period=' ~ period ~ '&from_date=' ~ from_date ~ '&to_date=' ~ to_date ~ '&sort=' ~ sort %}
    <nav class="mt-3">
      <ul class="pagination pagination-sm justify-content-center flex-wrap">
        {% if page > 1 %}<li class="page-item"><a class="page-link" href="?{{ qs }}&page={{ page-1 }}">‹</a></li>{% endif %}
        {% for p in range([1, page-3]|max, [total_pages+1, page+4]|min) %}
        <li class="page-item {% if p == page %}active{% endif %}"><a class="page-link" href="?{{ qs }}&page={{ p }}">{{ p }}</a></li>
        {% endfor %}
        {% if page < total_pages %}<li class="page-item"><a class="page-link" href="?{{ qs }}&page={{ page+1 }}">›</a></li>{% endif %}
      </ul>
      <p class="text-center mt-1" style="font-size:0.75rem;color:#3d4a5c;">стор. {{ page }} з {{ total_pages }}</p>
    </nav>
    {% endif %}

    {% else %}
    <div class="empty-state">
      <div style="font-size:2rem;margin-bottom:.5rem">{{ '🔍' if q or channel else '📭' }}</div>
      {{ 'Нічого не знайдено' if q or channel else 'Постів ще немає' }}
    </div>
    {% endif %}
  </div>


</div>

</div>
<!-- /report-root -->

<!-- Stop-word modal -->
<div id="stop-modal" class="stop-modal-overlay">
  <div class="stop-modal-box">
    <div class="stop-modal-title">Додати до стоп-слів?</div>
    <div class="stop-modal-desc">Слово <strong id="stop-modal-word"></strong> буде виключено з топу назавжди.</div>
    <div class="stop-modal-actions">
      <button class="stop-modal-no"  onclick="confirmStopWord(false)">Ні</button>
      <button class="stop-modal-yes" onclick="confirmStopWord(true)">Так</button>
    </div>
  </div>
</div>

<script>
function toggleCustomDates() {
  const block = document.getElementById('custom-dates');
  const opened = block.classList.toggle('show');
  // При відкритті custom-блоку фіксуємо period=custom, щоб submit "Застосувати"
  // не передавав попереднє значення (24h/7d тощо).
  if (opened) {
    const hidden = document.getElementById('period-hidden');
    if (hidden) hidden.value = 'custom';
    document.querySelectorAll('.preset-chip').forEach(b => b.classList.remove('active'));
    const customChip = document.getElementById('period-chip-custom');
    if (customChip) customChip.classList.add('active');
  }
}

// Чіпи періоду: змінюють приховане поле і одразу сабмітять форму. Через це
// кнопки "Знайти" / "Застосувати" зберігають поточний period замість дефолтного.
document.addEventListener('DOMContentLoaded', () => {
  const hidden = document.getElementById('period-hidden');
  const form   = document.getElementById('search-form');
  document.querySelectorAll('.period-chip[data-period]').forEach(btn => {
    btn.addEventListener('click', () => {
      if (!hidden || !form) return;
      hidden.value = btn.dataset.period;
      form.submit();
    });
  });
});

let _currentPeriod = {{ words_period|tojson }};
const _TOPIC_CATS = ['per','loc','org','term'];
const _CHANNEL_LIST = {{ channel_list | tojson }};

function escapeHtml(s) {
  return String(s).replace(/[&<>"']/g, c => ({
    '&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'
  }[c]));
}

function renderCategoryRows(rows) {
  if (!rows || !rows.length) {
    return '<div class="topic-empty">поки немає даних</div>';
  }
  const maxScore = rows[0][2] > 0 ? rows[0][2] : 1;
  return rows.map(([word, cnt, score]) => {
    const w = escapeHtml(word);
    const wJson = escapeHtml(JSON.stringify(word));
    return `
      <div class="word-row">
        <div class="d-flex justify-content-between align-items-center">
          <span class="d-flex align-items-center gap-2" style="min-width:0">
            <a href="/?q=${encodeURIComponent(word)}" class="word-row-link">${w}</a>
            <button class="stop-btn" title="Додати до стоп-слів" onclick='addStopWord(event, ${wJson})'>✕</button>
          </span>
          <span class="word-row-cnt" title="кількість згадок">${cnt}</span>
        </div>
        <div class="word-bar-track">
          <div class="word-bar-fill" style="width:${Math.round(score/maxScore*100)}%"></div>
        </div>
      </div>`;
  }).join('');
}

let _topWordsRetryTimer = null;

function loadTopWords(period) {
  _currentPeriod = period;
  if (_topWordsRetryTimer) { clearTimeout(_topWordsRetryTimer); _topWordsRetryTimer = null; }
  document.querySelectorAll('.period-tab[data-period]').forEach(btn => {
    btn.classList.toggle('active', btn.dataset.period === period);
  });
  _TOPIC_CATS.forEach(cat => {
    const list = document.getElementById('topic-list-' + cat);
    if (list) list.style.opacity = '0.4';
  });
  const topicsSection = document.querySelector('.topics-section');
  const ch = (topicsSection && topicsSection.dataset.channel) || '';
  const params = new URLSearchParams({ period });
  if (ch) params.set('channel', ch);
  fetch('/api/top-words?' + params.toString())
    .then(r => r.json())
    .then(data => {
      let totalRows = 0;
      _TOPIC_CATS.forEach(cat => {
        const list  = document.getElementById('topic-list-'  + cat);
        const count = document.getElementById('topic-count-' + cat);
        const rows  = data[cat] || [];
        totalRows += rows.length;
        if (list) {
          list.innerHTML = rows.length
            ? renderCategoryRows(rows)
            : '<div class="topic-empty">обчислюється у фоні…</div>';
          list.style.opacity = '1';
        }
        if (count) { count.textContent = rows.length; }
      });
      // Якщо все порожнє — кеш ще не готовий, повторити через 15с.
      if (totalRows === 0 && _currentPeriod === period) {
        _topWordsRetryTimer = setTimeout(() => loadTopWords(period), 15000);
      }
    })
    .catch(() => {
      _TOPIC_CATS.forEach(cat => {
        const list = document.getElementById('topic-list-' + cat);
        if (list) list.style.opacity = '1';
      });
    });
}

function addStopWord(e, word) {
  e.preventDefault();
  e.stopPropagation();
  const modal = document.getElementById('stop-modal');
  document.getElementById('stop-modal-word').textContent = '«' + word + '»';
  modal.dataset.word = word;
  modal.classList.add('show');
}

function confirmStopWord(yes) {
  const modal = document.getElementById('stop-modal');
  modal.classList.remove('show');
  if (!yes) return;
  const word = modal.dataset.word;
  fetch('/api/add-stop-word', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({word})
  }).then(() => loadTopWords(_currentPeriod));
}

// Старт: підвантажити топ-слова (якщо server-render віддав порожньо — загорнеться у retry-loop)
document.addEventListener('DOMContentLoaded', () => {
  loadTopWords(_currentPeriod);
});

// ── Динаміка згадок (Chart.js) ───────────────────────────────────────────────
let timelineChart = null;
const _MONTHS_UA = ['січ','лют','бер','кві','тра','чер','лип','сер','вер','жов','лис','гру'];

function fmtDateUA(iso) {                       // '2026-04-27' → '27 кві 2026'
  if (!iso) return '';
  const [y, m, d] = iso.split('-');
  return parseInt(d,10) + ' ' + _MONTHS_UA[parseInt(m,10)-1] + ' ' + y;
}

function pluralUA(n) {                          // 1 згадка / 2 згадки / 5 згадок
  const mod10  = n % 10;
  const mod100 = n % 100;
  if (mod10 === 1 && mod100 !== 11) return 'згадка';
  if (mod10 >= 2 && mod10 <= 4 && (mod100 < 12 || mod100 > 14)) return 'згадки';
  return 'згадок';
}

function fmtBig(n) {                            // 12345 → '12.3k', 1234567 → '1.2M'
  if (n == null) return '';
  if (n < 1000) return String(n);
  if (n < 1_000_000) return (n/1000).toFixed(n < 10000 ? 1 : 0) + 'k';
  return (n/1_000_000).toFixed(n < 10_000_000 ? 1 : 0) + 'M';
}

function fmtBigUA(n) {                          // 20500000 → '20.5 млн', 12300 → '12.3 тис.'
  if (n == null || !n) return '0';
  if (n < 1000) return String(n);
  if (n < 1_000_000) return (n/1000).toFixed(n < 10000 ? 1 : 0) + ' тис.';
  if (n < 1_000_000_000) return (n/1_000_000).toFixed(n < 10_000_000 ? 1 : 0) + ' млн';
  return (n/1_000_000_000).toFixed(n < 10_000_000_000 ? 1 : 0) + ' млрд';
}

let _reachPollTimer = null;
let _reachCurrentDays = null;       // щоб старий polling не «затирав» свіжий графік

function setReachStatus(text, cls) {
  const el = document.getElementById('reach-status');
  if (!el) return;
  if (!text) { el.style.display = 'none'; el.textContent = ''; el.className = 'reach-status'; return; }
  el.textContent = text;
  el.className = 'reach-status' + (cls ? ' ' + cls : '');
  el.style.display = 'block';
}

function applyReachData(reachData, days) {
  if (!timelineChart || _reachCurrentDays !== days) return;
  const byDate = {};
  reachData.forEach(d => { byDate[d.date] = d.views; });
  const values = timelineChart.data.labels.map(lbl => byDate[lbl] ?? 0);
  // Видаляємо стару reach-лінію якщо була
  timelineChart.data.datasets = timelineChart.data.datasets.filter(ds => ds._kind !== 'reach');
  timelineChart.data.datasets.push({
    _kind: 'reach',
    label: 'Охоплення',
    data: values,
    yAxisID: 'yReach',
    borderColor: '#56cfe1',
    backgroundColor: 'rgba(86,207,225,0.08)',
    tension: 0.3,
    fill: false,
    pointRadius: 2,
    pointHoverRadius: 5,
    pointHoverBackgroundColor: '#90e0ef',
    borderWidth: 2,
    borderDash: [4, 3],
  });
  timelineChart.options.scales.yReach = {
    position: 'right',
    beginAtZero: true,
    ticks: { color: '#56cfe1', callback: (v) => fmtBig(v) },
    grid: { drawOnChartArea: false }
  };
  timelineChart.options.plugins.legend = {
    display: true,
    labels: { color: '#6b7280', boxWidth: 14, font: { size: 11 } }
  };
  // Підпис «Згадки» для першого датасету
  if (timelineChart.data.datasets[0] && !timelineChart.data.datasets[0].label) {
    timelineChart.data.datasets[0].label = 'Згадки';
  }
  // Сумарне охоплення поряд з заголовком
  const totalViews = values.reduce((s, v) => s + (v || 0), 0);
  const totalEl = document.getElementById('timeline-reach-total');
  if (totalEl) {
    totalEl.textContent = totalViews ? `Σ охоплення: ${fmtBigUA(totalViews)} переглядів` : '';
  }
  timelineChart.update();
}

async function loadReach(days) {
  if (_reachPollTimer) { clearTimeout(_reachPollTimer); _reachPollTimer = null; }
  const card = document.getElementById('timeline-card');
  if (!card) return;
  const q       = card.dataset.q || '';
  const channel = card.dataset.channel || '';
  if (!q) { setReachStatus(''); return; }       // без q охоплення не рахуємо

  if (![7, 30].includes(days)) {
    setReachStatus('Охоплення доступне для періодів 7 і 30 днів', '');
    return;
  }

  setReachStatus('⏳ запит охоплення…');

  let resp;
  try {
    resp = await fetch('/api/timeline-reach', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ q, channel, days }),
    }).then(r => r.json().then(j => ({ status: r.status, body: j })));
  } catch (e) {
    setReachStatus('Помилка мережі при запиті охоплення', 'error');
    return;
  }

  if (resp.status === 429 && resp.body.error === 'queue_full') {
    setReachStatus('⏸ занадто багато активних запитів охоплення, спробуйте пізніше', 'error');
    return;
  }
  if (resp.body.error === 'q_required') {
    setReachStatus('');
    return;
  }
  if (resp.status !== 200) {
    setReachStatus('Помилка: ' + (resp.body.error || resp.status), 'error');
    return;
  }

  if (resp.body.state === 'done') {
    applyReachData(resp.body.data, days);
    setReachStatus('Охоплення (з кешу): сума переглядів по днях', 'done');
    return;
  }

  // Якщо одразу стали в чергу — показуємо позицію.
  if (resp.body.state === 'queued') {
    setReachStatus('⏳ у черзі на підрахунок охоплення: май витримку…');
  }

  // Polling
  const taskId = resp.body.task_id;
  const poll = async () => {
    if (_reachCurrentDays !== days) return;     // перемкнули період → стоп
    let st;
    try {
      st = await fetch('/api/timeline-reach/' + taskId).then(r => r.json());
    } catch (e) {
      setReachStatus('Втрачено зв’язок з task ' + taskId, 'error');
      return;
    }
    if (st.state === 'done') {
      applyReachData(st.data, days);
      const totalViews = (st.data || []).reduce((sum, d) => sum + (d.views || 0), 0);
      setReachStatus(`Охоплення зібрано та становить ${fmtBigUA(totalViews)} переглядів`, 'done');
      return;
    }
    if (st.state === 'error') {
      const msg = st.error || 'unknown';
      setReachStatus('Не вдалось зібрати охоплення: ' + msg, 'error');
      return;
    }
    if (st.state === 'queued') {
      setReachStatus('⏳ у черзі на підрахунок охоплення: май витримку…');
      _reachPollTimer = setTimeout(poll, 2000);
      return;
    }
    const total = st.total || 0;
    const prog  = st.progress || 0;
    setReachStatus(`⏳ збираємо охоплення (${prog}/${total} каналів)…`);
    _reachPollTimer = setTimeout(poll, 2000);
  };
  _reachPollTimer = setTimeout(poll, 1500);
}

// Підписи значень безпосередньо на графіку.
// • Згадки (перший dataset, count) — над точкою.
// • Охоплення (dataset з _kind='reach', views) — під точкою, у форматі fmtBigUA.
// Малюємо локальні максимуми + крайні точки завжди, інші ненульові — якщо вистачає місця.
const pointValueLabelsPlugin = {
  id: 'pointValueLabels',
  afterDatasetsDraw(chart) {
    const {ctx} = chart;
    const datasets = chart.data.datasets || [];
    ctx.save();
    ctx.font = '600 10px system-ui, sans-serif';
    ctx.textAlign = 'center';

    datasets.forEach((ds, dsIdx) => {
      if (!ds || !ds.data || !ds.data.length) return;
      const meta = chart.getDatasetMeta(dsIdx);
      if (!meta || !meta.data) return;
      const data = ds.data;
      const last = data.length - 1;
      const isReach = ds._kind === 'reach';
      const isPeak = (i) => {
        const v = data[i];
        if (!v) return false;
        const l = i > 0 ? data[i-1] : -Infinity;
        const r = i < last ? data[i+1] : -Infinity;
        return v >= l && v >= r;
      };
      const fmt = isReach ? fmtBigUA : (v) => String(v);
      const isPdf = document.body.classList.contains('pdf-mode');
      ctx.fillStyle = isPdf
        ? (isReach ? '#0c7588' : '#3a2db8')
        : (isReach ? '#90e0ef' : '#c4bfff');
      ctx.textBaseline = isReach ? 'top' : 'bottom';
      const minStep = isReach ? 38 : 28;
      let lastDrawX = -Infinity;
      for (let i = 0; i < data.length; i++) {
        const v = data[i];
        if (!v) continue;
        const point = meta.data[i];
        if (!point) continue;
        const force = (i === 0) || (i === last) || isPeak(i);
        if (!force && point.x - lastDrawX < minStep) continue;
        const yOffset = isReach ? 6 : -6;
        ctx.fillText(fmt(v), point.x, point.y + yOffset);
        lastDrawX = point.x;
      }
    });
    ctx.restore();
  }
};

async function loadTimeline(days) {
  if (typeof Chart === 'undefined') return;     // CDN ще не завантажений
  const card = document.getElementById('timeline-card');
  if (!card) return;
  _reachCurrentDays = days;
  // Скидаємо попередню Σ охоплення при зміні діапазону днів
  const totalEl = document.getElementById('timeline-reach-total');
  if (totalEl) totalEl.textContent = '';
  const q       = card.dataset.q || '';
  const channel = card.dataset.channel || '';
  const params  = new URLSearchParams({ q, channel, days: String(days) });
  const data    = await fetch('/api/timeline?' + params).then(r => r.json());

  const labels = data.map(d => d.date);
  const values = data.map(d => d.count);

  if (timelineChart) timelineChart.destroy();
  timelineChart = new Chart(document.getElementById('timeline-chart'), {
    type: 'line',
    plugins: [pointValueLabelsPlugin],
    data: {
      labels,
      datasets: [{
        label: 'Згадки',
        data: values,
        yAxisID: 'y',
        borderColor: '#8b83ff',
        backgroundColor: 'rgba(139,131,255,0.15)',
        tension: 0.3,
        fill: true,
        pointRadius: 2,
        pointHoverRadius: 5,
        pointHoverBackgroundColor: '#c4bfff',
        borderWidth: 2,
      }]
    },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      interaction: { mode: 'index', intersect: false },
      plugins: {
        legend: { display: false },
        tooltip: {
          backgroundColor: 'rgba(13,17,28,0.95)',
          borderColor: 'rgba(108,99,255,0.4)', borderWidth: 1,
          titleColor: '#c4bfff', bodyColor: '#e2e8f0',
          padding: 10, displayColors: false,
          callbacks: {
            title: (items) => fmtDateUA(items[0].label),
            label: (ctx) => {
              const isReach = ctx.dataset._kind === 'reach';
              if (isReach) return 'Охоплення: ' + fmtBig(ctx.parsed.y) + ' переглядів';
              return ctx.parsed.y + ' ' + pluralUA(ctx.parsed.y);
            }
          }
        }
      },
      scales: {
        x: {
          ticks: {
            color: '#4a5568',
            maxTicksLimit: days <= 7 ? 7 : (days <= 30 ? 10 : 12),
            callback: function(val) {              // короткий формат: '27.04'
              const iso = this.getLabelForValue(val);
              return iso ? iso.slice(5).replace('-', '.') : '';
            }
          },
          grid: { color: 'rgba(255,255,255,0.04)' }
        },
        y: {
          position: 'left',
          beginAtZero: true,
          ticks: { color: '#4a5568', precision: 0 },
          grid: { color: 'rgba(255,255,255,0.04)' }
        }
      }
    }
  });

  // Охоплення підвантажуємо асинхронно; працює лише за наявності q.
  if (q) {
    loadReach(days);
  } else {
    setReachStatus('');
  }
}

// ── Завантаження xlsx: візуальний фідбек і захист від подвійного кліку ──────
// Експорт використовує query-параметри поточної сторінки (а не серіалізацію форми),
// щоб гарантовано завантажити саме ту вибірку, яку бачить користувач. Інакше
// hidden `period=custom` всередині custom-dates перебивав активний preset.
document.addEventListener('DOMContentLoaded', () => {
  const dlBtn = document.getElementById('export-xlsx-btn');
  if (!dlBtn) return;
  const originalText = dlBtn.textContent;
  let inFlight = false;

  function reset() {
    inFlight = false;
    dlBtn.disabled = false;
    dlBtn.textContent = originalText;
    dlBtn.style.opacity = '';
  }

  dlBtn.addEventListener('click', () => {
    if (inFlight) return;
    inFlight = true;
    dlBtn.dataset.startedAt = String(performance.now());
    dlBtn.disabled = true;
    dlBtn.style.opacity = '0.6';
    dlBtn.textContent = '⏳ Збирається... (може зайняти кілька хвилин)';
    window.location.href = '/api/export-xlsx' + window.location.search;
  });

  // Коли користувач повертається на вкладку (після save dialog) — повертаємо кнопку
  window.addEventListener('focus', () => { if (inFlight) reset(); });
  window.addEventListener('pageshow', () => { if (inFlight) reset(); });
  // Гарантований fallback через 15 хв
  setInterval(() => {
    if (inFlight && performance.now() - Number(dlBtn.dataset.startedAt || 0) > 900000) reset();
  }, 30000);
});

// ── AI-аналітика: запит до /api/analytics + рендер замість стрічки постів ────
document.addEventListener('DOMContentLoaded', () => {
  const btn = document.getElementById('analytics-btn');
  if (!btn || btn.disabled) return;
  btn.addEventListener('click', startAnalytics);
});

const ANALYTICS_ERR = {
  q_required:         'Спочатку введіть ключове слово в полі пошуку.',
  period_not_allowed: 'AI-аналітика доступна лише для періодів «Доба» або «Тиждень».',
  no_posts:           'За цим запитом немає постів у вибраному періоді.',
  empty_texts:        'Знайдені пости порожні — нічого аналізувати.',
  key_missing:        'На сервері немає файлу /root/.anthropic_key — налаштуйте API-ключ Claude.',
  key_empty:          'Файл /root/.anthropic_key порожній.',
  sdk_missing:        'На сервері не встановлено пакет anthropic — потрібен pip install anthropic.',
};

async function startAnalytics() {
  const btn = document.getElementById('analytics-btn');
  const orig = btn.textContent;
  btn.disabled = true;
  btn.textContent = '⏳ Аналізую...';

  // Параметри беремо з поточного URL — рівно ті ж фільтри, що й на сторінці
  const u = new URL(window.location.href);
  const params = {
    q:         u.searchParams.get('q') || '',
    channel:   u.searchParams.get('channel') || '',
    period:    u.searchParams.get('period') || '24h',
    from_date: u.searchParams.get('from_date') || '',
    to_date:   u.searchParams.get('to_date') || '',
  };

  try {
    const r = await fetch('/api/analytics', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(params),
    });
    if (r.status === 429) {
      alert('Зайнято — інша AI-аналітика виконується. Спробуйте за 5 секунд.');
      return;
    }
    const j = await r.json();
    if (j.error) { alert(ANALYTICS_ERR[j.error] || ('Помилка: ' + j.error)); return; }
    if (j.state === 'done') return renderAnalytics(j.data, params.q);

    while (true) {
      await new Promise(res => setTimeout(res, 2500));
      const s = await fetch('/api/analytics/' + j.task_id).then(x => x.json());
      if (s.state === 'done')  return renderAnalytics(s.data, params.q);
      if (s.state === 'error') {
        alert(ANALYTICS_ERR[s.error] || ('Помилка: ' + s.error));
        return;
      }
    }
  } catch (e) {
    alert('Мережева помилка: ' + e.message);
  } finally {
    btn.disabled = false;
    btn.textContent = orig;
  }
}

function renderAnalytics(data, q) {
  document.getElementById('analytics-q').textContent = `«${q}»`;
  const used = data.posts_used, total = data.posts_total;
  let meta = `Проаналізовано ${used}`;
  if (used < total) meta += ` з ${total} (обрізано через ліміт контексту)`;
  else              meta += ` постів`;
  meta += ` · модель ${data.model}`;
  if (data.elapsed_sec) meta += ` · ${data.elapsed_sec}с`;
  document.getElementById('analytics-meta').textContent = meta;
  document.getElementById('analytics-body').innerHTML = data.html;
  document.getElementById('analytics-result').style.display = 'block';
  // Згідно ТЗ — стрічка постів ховається на час перегляду аналітики
  const grid = document.getElementById('main-grid');
  if (grid) grid.style.display = 'none';
  document.getElementById('analytics-result').scrollIntoView({behavior: 'smooth', block: 'start'});
}

function closeAnalytics() {
  document.getElementById('analytics-result').style.display = 'none';
  const grid = document.getElementById('main-grid');
  if (grid) grid.style.display = '';
}

document.addEventListener('DOMContentLoaded', () => {
  const tabs = document.querySelectorAll('.timeline-tab');
  if (!tabs.length) return;

  // Вибраний період зберігаємо у localStorage, щоб після Refresh або «Знайти»
  // не стрибало на дефолт. Default — 7 днів.
  const ALLOWED = [7, 30, 90];
  const LS_KEY = 'mention_days';
  let savedDays = parseInt(localStorage.getItem(LS_KEY), 10);
  if (!ALLOWED.includes(savedDays)) savedDays = 7;
  const setActive = (days) => {
    tabs.forEach(b => b.classList.toggle('active',
      parseInt(b.dataset.days, 10) === days));
  };
  setActive(savedDays);

  tabs.forEach(btn => {
    btn.addEventListener('click', () => {
      const days = parseInt(btn.dataset.days, 10);
      try { localStorage.setItem(LS_KEY, String(days)); } catch (e) {}
      setActive(days);
      loadTimeline(days);
    });
  });
  // Chart.js підвантажується з defer — даємо йому з'явитися
  if (typeof Chart === 'undefined') {
    window.addEventListener('load', () => loadTimeline(savedDays));
  } else {
    loadTimeline(savedDays);
  }
});

// ── Channel autocomplete ─────────────────────────────────────────────────────
(function() {
  const wrap   = document.getElementById('channel-ac');
  if (!wrap) return;
  const input  = document.getElementById('channel-ac-input');
  const list   = document.getElementById('channel-ac-list');
  const clear  = document.getElementById('channel-ac-clear');
  const MAX_VISIBLE = 50;
  let activeIdx = -1;
  let visibleItems = [];

  function normalize(s) { return String(s || '').toLowerCase(); }

  function highlight(name, query) {
    const n = escapeHtml(name);
    if (!query) return n;
    const q = normalize(query);
    const lo = normalize(name);
    const i = lo.indexOf(q);
    if (i < 0) return n;
    const before = escapeHtml(name.slice(0, i));
    const match  = escapeHtml(name.slice(i, i + query.length));
    const after  = escapeHtml(name.slice(i + query.length));
    return before + '<mark>' + match + '</mark>' + after;
  }

  function render(query) {
    const q = normalize(query);
    let items;
    if (!q) {
      items = _CHANNEL_LIST.slice(0, MAX_VISIBLE);
    } else {
      items = [];
      for (const ch of _CHANNEL_LIST) {
        if (normalize(ch).includes(q)) {
          items.push(ch);
          if (items.length >= MAX_VISIBLE) break;
        }
      }
    }
    visibleItems = items;
    activeIdx = -1;
    if (!items.length) {
      list.innerHTML = '<div class="channel-ac-empty">нічого не знайдено</div>';
      return;
    }
    list.innerHTML = items.map((ch, i) =>
      '<div class="channel-ac-item" data-idx="' + i + '">' + highlight(ch, query) + '</div>'
    ).join('');
  }

  function open() { render(input.value); wrap.classList.add('open'); }
  function close() { wrap.classList.remove('open'); activeIdx = -1; }

  function setActive(i) {
    const items = list.querySelectorAll('.channel-ac-item');
    if (!items.length) return;
    if (i < 0) i = items.length - 1;
    if (i >= items.length) i = 0;
    activeIdx = i;
    items.forEach((el, idx) => el.classList.toggle('active', idx === i));
    items[i].scrollIntoView({ block: 'nearest' });
  }

  function pick(value) {
    input.value = value;
    wrap.classList.toggle('has-value', !!value);
    close();
    // одразу шукати по обраному каналу
    input.form && input.form.submit();
  }

  input.addEventListener('focus', open);
  input.addEventListener('input', () => {
    wrap.classList.toggle('has-value', !!input.value);
    open();
  });
  input.addEventListener('keydown', (e) => {
    if (e.key === 'ArrowDown') { e.preventDefault(); if (!wrap.classList.contains('open')) open(); setActive(activeIdx + 1); }
    else if (e.key === 'ArrowUp') { e.preventDefault(); setActive(activeIdx - 1); }
    else if (e.key === 'Enter') {
      if (activeIdx >= 0 && visibleItems[activeIdx]) {
        e.preventDefault();
        pick(visibleItems[activeIdx]);
      } else if (input.value.trim() && visibleItems.length) {
        // користувач набрав текст і натиснув Enter без явного вибору —
        // беремо найрелевантніший варіант зі списку (substring-match)
        e.preventDefault();
        pick(visibleItems[0]);
      }
      // інакше (порожнє поле) — даємо формі засабмітитись як є
    } else if (e.key === 'Escape') {
      close();
    }
  });

  list.addEventListener('mousedown', (e) => {
    const item = e.target.closest('.channel-ac-item');
    if (!item) return;
    e.preventDefault();
    const i = parseInt(item.dataset.idx, 10);
    if (visibleItems[i]) pick(visibleItems[i]);
  });

  clear.addEventListener('click', () => {
    input.value = '';
    wrap.classList.remove('has-value');
    input.form && input.form.submit();
  });

  document.addEventListener('click', (e) => {
    if (!wrap.contains(e.target)) close();
  });
})();

// ── PDF звіт (html2canvas → jsPDF, одна сторінка з заголовком) ──────────────
async function generatePDF(btn) {
  if (!window.html2canvas || !window.jspdf) {
    alert('Бібліотеки PDF ще завантажуються — спробуй за мить.');
    return;
  }
  const root = document.getElementById('report-root');
  if (!root) return;

  const origText = btn.textContent;
  btn.disabled = true;
  btn.textContent = '⏳ Готую…';

  // Витягуємо параметри з URL для заголовку
  const u = new URL(window.location.href);
  const q = u.searchParams.get('q') || '';
  const periodLabel = {{ period_label|tojson }};
  const periodDates = {{ period_dates|tojson }};

  // Тимчасовий заголовок — вставляємо першим дочірнім у report-root
  const header = document.createElement('div');
  header.id = 'pdf-report-header';
  header.style.cssText =
    'background:#ffffff;color:#000;padding:1.5rem 2rem 1rem;margin:0 auto 1rem;' +
    'max-width:var(--page-max);' +
    'border-bottom:2px solid #6c63ff;text-align:center;' +
    'font-family:system-ui,sans-serif;';
  const subParts = [];
  if (q) subParts.push(`по ключовому слову «${q}»`);
  if (periodLabel) subParts.push(periodLabel + (periodDates ? ` (${periodDates})` : ''));
  header.innerHTML =
    '<div style="font-size:1.3rem;font-weight:800;color:#000;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:.4rem;">' +
      'МОНІТОРИНГ МЕДІА-ПРОСТОРУ МЕСЕНДЖЕРА «MAX»' +
    '</div>' +
    '<div style="font-size:1rem;font-weight:600;color:#000;">' +
      'Звіт ' + (subParts.length ? subParts.join(' · ') : '— загальний') +
    '</div>';
  root.insertBefore(header, root.firstChild);

  // Перемикаємо у світлу тему для PDF
  document.body.classList.add('pdf-mode');

  // Патч Chart.js: темні axis-кольори → чорні, grid-лінії → світло-сірі
  let chartPatch = null;
  if (typeof timelineChart !== 'undefined' && timelineChart) {
    const sx = timelineChart.options.scales.x;
    const sy = timelineChart.options.scales.y;
    const sr = timelineChart.options.scales.yReach;
    const lg = timelineChart.options.plugins && timelineChart.options.plugins.legend;
    chartPatch = {
      x_t: sx.ticks.color, x_g: sx.grid.color,
      y_t: sy.ticks.color, y_g: sy.grid.color,
      r_t: sr ? sr.ticks.color : null,
      l_c: lg && lg.labels ? lg.labels.color : null,
    };
    sx.ticks.color = '#000'; sx.grid.color = '#dddddd';
    sy.ticks.color = '#000'; sy.grid.color = '#dddddd';
    if (sr) sr.ticks.color = '#000';
    if (lg && lg.labels) lg.labels.color = '#000';
    timelineChart.update('none');
    // Контейнер графіка щойно розширився через pdf-mode CSS — Chart.js про це
    // не дізнається до наступного resize-обсервера, тому форсуємо перерахунок
    // розмірів canvas і чекаємо два кадри щоб браузер відмалював усе у нову ширину.
    timelineChart.resize();
    await new Promise(r => requestAnimationFrame(r));
    await new Promise(r => requestAnimationFrame(r));
  }

  try {
    const canvas = await html2canvas(root, {
      scale: 2,
      backgroundColor: '#ffffff',
      useCORS: true,
      logging: false,
      windowWidth: root.scrollWidth,
      ignoreElements: (el) => {
        if (!el.classList) return false;
        return el.classList.contains('stat-grid')          // 4 картки зверху (Всього постів / Каналів / За годину / За 24 год)
            || el.classList.contains('search-wrap')
            || el.classList.contains('channel-ac-list')
            || el.classList.contains('stop-btn')
            || el.classList.contains('stop-modal-overlay');
      }
    });

    const { jsPDF } = window.jspdf;
    const pdf = new jsPDF({ orientation: 'p', unit: 'mm', format: 'a4' });
    const pageW = pdf.internal.pageSize.getWidth();   // 210
    const pageH = pdf.internal.pageSize.getHeight();  // 297
    const margin = 15;                                 // стандартні відступи для друку (мм)
    const drawW = pageW - margin * 2;                  // ширина блоку = ширина листа − поля
    const availH = pageH - margin * 2;                 // макс. висота на одну сторінку

    // Розтягуємо знімок по ширині A4 (мінус поля). Якщо висота не вміщується —
    // ріжемо канвас на горизонтальні смуги і кладемо їх на послідовні сторінки.
    const pxPerMm = canvas.width / drawW;
    const sliceHpx = Math.floor(availH * pxPerMm);

    let sy = 0;
    let pageIdx = 0;
    while (sy < canvas.height) {
      const h = Math.min(sliceHpx, canvas.height - sy);
      const slice = document.createElement('canvas');
      slice.width = canvas.width;
      slice.height = h;
      const sctx = slice.getContext('2d');
      sctx.fillStyle = '#ffffff';
      sctx.fillRect(0, 0, canvas.width, h);
      sctx.drawImage(canvas, 0, sy, canvas.width, h, 0, 0, canvas.width, h);
      const sliceData = slice.toDataURL('image/jpeg', 0.92);
      if (pageIdx > 0) pdf.addPage();
      pdf.addImage(sliceData, 'JPEG', margin, margin, drawW, h / pxPerMm);
      sy += h;
      pageIdx++;
    }

    const d = new Date();
    const pad = n => String(n).padStart(2,'0');
    const ts = `${d.getFullYear()}-${pad(d.getMonth()+1)}-${pad(d.getDate())}_${pad(d.getHours())}${pad(d.getMinutes())}`;
    pdf.save(`max_report_${ts}.pdf`);
  } catch (err) {
    console.error('[report] failed:', err);
    alert('Не вдалось сформувати звіт: ' + (err && err.message || err));
  } finally {
    // Прибираємо тимчасовий заголовок — у будь-якому разі
    const h = document.getElementById('pdf-report-header');
    if (h) h.remove();
    // Знімаємо PDF-режим
    document.body.classList.remove('pdf-mode');
    // Відновлюємо кольори графіка
    if (chartPatch && typeof timelineChart !== 'undefined' && timelineChart) {
      const sx = timelineChart.options.scales.x;
      const sy = timelineChart.options.scales.y;
      const sr = timelineChart.options.scales.yReach;
      const lg = timelineChart.options.plugins && timelineChart.options.plugins.legend;
      sx.ticks.color = chartPatch.x_t; sx.grid.color = chartPatch.x_g;
      sy.ticks.color = chartPatch.y_t; sy.grid.color = chartPatch.y_g;
      if (sr && chartPatch.r_t) sr.ticks.color = chartPatch.r_t;
      if (lg && lg.labels && chartPatch.l_c) lg.labels.color = chartPatch.l_c;
      timelineChart.update('none');
    }
    btn.disabled = false;
    btn.textContent = origText;
  }
}
</script>
</body>
</html>
"""

def get_db():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA cache_size=-32000")
    return conn


# ── FTS5 full-text search ─────────────────────────────────────────────────────
# Віртуальна таблиця з unicode61 tokenizer (case-insensitive для кирилиці)
# + тригери на INSERT/UPDATE/DELETE у messages для автосинхронізації.
# Парсер нічого не знає про FTS — пише у messages як раніше.

def init_fts(db: sqlite3.Connection):
    """
    Ідемпотентна ініціалізація FTS5:
    1. Створює віртуальну таблицю з unicode61 tokenizer (IF NOT EXISTS)
    2. Створює тригери INSERT/UPDATE/DELETE для синхронізації з `messages`
    3. Перевіряє заповненість індексу через probe-запит (частотні літери)
       і робить 'rebuild' якщо видно що індекс порожній або неповний.

    Для FTS5 з external content (content='messages') команда 'rebuild' —
    єдиний правильний спосіб пересинхронізувати індекс з content table.
    INSERT INTO messages_fts SELECT для external content НЕ індексує текст,
    а тільки реєструє rowid'и.
    """
    db.execute("""
        CREATE VIRTUAL TABLE IF NOT EXISTS messages_fts USING fts5(
            text,
            channel_title UNINDEXED,
            content='messages',
            content_rowid='id',
            tokenize = "unicode61 remove_diacritics 2"
        )
    """)
    db.execute("""
        CREATE TRIGGER IF NOT EXISTS messages_ai AFTER INSERT ON messages BEGIN
            INSERT INTO messages_fts(rowid, text, channel_title)
                VALUES (new.id, new.text, new.channel_title);
        END
    """)
    db.execute("""
        CREATE TRIGGER IF NOT EXISTS messages_ad AFTER DELETE ON messages BEGIN
            INSERT INTO messages_fts(messages_fts, rowid, text, channel_title)
                VALUES('delete', old.id, old.text, old.channel_title);
        END
    """)
    db.execute("""
        CREATE TRIGGER IF NOT EXISTS messages_au AFTER UPDATE ON messages BEGIN
            INSERT INTO messages_fts(messages_fts, rowid, text, channel_title)
                VALUES('delete', old.id, old.text, old.channel_title);
            INSERT INTO messages_fts(rowid, text, channel_title)
                VALUES (new.id, new.text, new.channel_title);
        END
    """)

    # Probe-перевірка: шукаємо дуже частотні літери/склади російської мови.
    # Якщо FTS порожній або неповний — знайдемо мало.
    msg_count = db.execute("SELECT COUNT(*) FROM messages").fetchone()[0]
    if msg_count == 0:
        db.commit()
        return

    try:
        probe = db.execute(
            "SELECT COUNT(*) FROM messages_fts "
            "WHERE messages_fts MATCH 'а* OR о* OR е* OR и*'"
        ).fetchone()[0]
    except sqlite3.OperationalError:
        probe = 0

    # Якщо індекс явно неповний (<50% від повідомлень знайдено probe-запитом) → rebuild
    if probe < msg_count // 2:
        print(f"[fts] REBUILD: probe={probe} / messages={msg_count}", flush=True)
        db.execute("INSERT INTO messages_fts(messages_fts) VALUES('rebuild')")
        print(f"[fts] готово: {msg_count} повідомлень проіндексовано", flush=True)
    else:
        print(f"[fts] індекс консистентний: probe={probe} / messages={msg_count}", flush=True)

    db.commit()


# ── Часові фільтри ────────────────────────────────────────────────────────────

PERIOD_DELTAS = {
    "1h":  timedelta(hours=1),
    "24h": timedelta(days=1),
    "7d":  timedelta(days=7),
    "30d": timedelta(days=30),
}
PERIOD_LABELS = {
    "1h": "за останню годину", "24h": "за останні 24 години",
    "7d": "за останні 7 днів", "30d": "за останні 30 днів",
    "all": "за весь час",      "custom": "за вибраний період",
}
DEFAULT_PERIOD = "24h"
_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')


def _format_period_dates(period: str, since_ts: str | None, until_ts: str | None) -> str:
    """Повертає рядок з датами діапазону для UI: 'DD.MM.YYYY — DD.MM.YYYY'.
    Для period='all' повертає порожній рядок."""
    if period == "all":
        return ""
    def _fmt(ts: str) -> str:
        try:
            return datetime.strptime(ts[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
        except ValueError:
            return ts[:10]
    now_str = datetime.now().strftime("%d.%m.%Y")
    if since_ts and until_ts:
        return f"{_fmt(since_ts)} — {_fmt(until_ts)}"
    if since_ts:
        return f"{_fmt(since_ts)} — {now_str}"
    if until_ts:
        return f"… — {_fmt(until_ts)}"
    return ""


def parse_time_filter(period: str, from_date: str, to_date: str) -> tuple[str | None, str | None, str]:
    """
    Перетворює (period, from_date, to_date) на (since_ts, until_ts, normalized_period).
    Повертає None для відсутньої межі. Невалідні значення — fallback на 'all'.
    """
    if period in PERIOD_DELTAS:
        since = (datetime.now() - PERIOD_DELTAS[period]).strftime("%Y-%m-%d %H:%M:%S")
        return since, None, period
    if period == "custom":
        valid_from = from_date if _DATE_RE.match(from_date or "") else None
        valid_to   = to_date   if _DATE_RE.match(to_date   or "") else None
        if not valid_from and not valid_to:
            return None, None, "all"
        since = f"{valid_from} 00:00:00" if valid_from else None
        until = f"{valid_to} 23:59:59"   if valid_to   else None
        return since, until, "custom"
    return None, None, "all"


_FTS_OPERATORS_RE = re.compile(r'\b(AND|OR|NOT|NEAR)\b|["():]', re.IGNORECASE)
_FTS_TOKEN_RE     = re.compile(r'[\w*]+', re.UNICODE)

# Snowball-стемер для російської — rule-based, без зовнішніх ресурсів.
# Потрібен щоб "зеленский" -> "зеленск*" (а не "зеленский*"), бо префіксний пошук
# у FTS5 матчить тільки токени що починаються з префікса; форми "зеленского",
# "зеленскому" не починаються на "зеленский" і ловляться лише через stem.
try:
    from nltk.stem.snowball import SnowballStemmer
    _RU_STEMMER = SnowballStemmer("russian")
except Exception:
    _RU_STEMMER = None

# Мінімальна довжина stem-у. Якщо стемер видав щось коротше (рідко: дуже короткі
# службові слова) — використовуємо оригінальний токен, щоб не отримати "ид*" з "идти".
_MIN_STEM_LEN = 4


def _stem_for_search(token: str) -> str:
    """Stem російського токена для префіксного пошуку. Повертає оригінал,
    якщо стемер недоступний, токен короткий, або stem занадто короткий."""
    if _RU_STEMMER is None:
        return token
    if len(token) <= _MIN_STEM_LEN:
        return token
    if not any('а' <= c <= 'я' or c == 'ё' for c in token.lower()):
        return token  # не-кирилиця: цифри, латиниця, тощо — не стемимо
    try:
        s = _RU_STEMMER.stem(token.lower())
    except Exception:
        return token
    return s if len(s) >= _MIN_STEM_LEN else token


def build_fts_query(q: str) -> str:
    """
    Перетворює користувацький ввід у безпечний FTS5 query.
    - Якщо q містить FTS-оператори (AND/OR/NOT/NEAR, лапки, дужки) — передаємо as-is.
    - Інакше — розбиваємо на токени, проганяємо через Snowball-стемер і додаємо *
      для префіксного пошуку. «зеленский» → «зеленск*», ловить усі словоформи.
    - Якщо користувач уже поставив * — не стемимо і не дублюємо (явний намір).
    - AND-логіка між токенами (дефолт FTS5).
    """
    q = q.strip()
    if not q:
        return ""
    if _FTS_OPERATORS_RE.search(q):
        return q
    tokens = _FTS_TOKEN_RE.findall(q)
    result = []
    for t in tokens:
        if not t:
            continue
        if t.endswith("*"):
            result.append(t)
        else:
            stem = _stem_for_search(t)
            result.append(f"{stem}*")
    return " ".join(result)

_STATS_CACHE_TTL = 30
_stats_cache: tuple[float, dict] | None = None
_stats_lock = threading.Lock()

# DISTINCT channel_title і GROUP BY по messages — повний скан таблиці на кожен
# рендер `/`. Кешуємо: список каналів змінюється рідко, топ — теж не миттєво.
_CHANNEL_LIST_TTL = 300
_TOP_CHANNELS_TTL = 60
_channel_list_cache: tuple[float, list] | None = None
_channel_list_lock = threading.Lock()
_top_channels_cache: dict[tuple, tuple[float, list]] = {}
_top_channels_lock = threading.Lock()

# Alert-канали (БПЛА/тривоги/радари/ППО). Перечитуємо за mtime файлу,
# рестарт dashboard після правок channels/alert_channels.txt не потрібен.
_alert_channels_cache: tuple[float, set[str]] = (0.0, set())
_alert_channels_lock = threading.Lock()


def _load_alert_channels() -> set[str]:
    """Lower-case alias-и alert-каналів. Кеш інвалідується по mtime файлу."""
    global _alert_channels_cache
    try:
        mtime = ALERT_CHANNELS_FILE.stat().st_mtime
    except OSError:
        return set()

    with _alert_channels_lock:
        cached_mtime, cached_set = _alert_channels_cache
        if cached_mtime == mtime:
            return cached_set

    aliases: set[str] = set()
    try:
        with open(ALERT_CHANNELS_FILE, encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith("#"):
                    continue
                aliases.add(s.lower())
    except OSError:
        return set()

    with _alert_channels_lock:
        _alert_channels_cache = (mtime, aliases)
    return aliases


def get_channel_list(db) -> list[str]:
    global _channel_list_cache
    with _channel_list_lock:
        if _channel_list_cache and time.time() - _channel_list_cache[0] < _CHANNEL_LIST_TTL:
            return _channel_list_cache[1]
    rows = db.execute(
        "SELECT DISTINCT channel_title FROM messages ORDER BY channel_title"
    ).fetchall()
    data = [r[0] for r in rows]
    with _channel_list_lock:
        _channel_list_cache = (time.time(), data)
    return data


def get_top_channels(db, q: str, since_ts: str | None, until_ts: str | None,
                     mode: str = "all") -> list:
    """mode: 'all' — без розділення; 'main' — exclude alert; 'alert' — only alert."""
    alert_aliases = _load_alert_channels() if mode != "all" else set()
    if mode == "alert" and not alert_aliases:
        return []

    key = (q, since_ts, until_ts, mode)
    with _top_channels_lock:
        cached = _top_channels_cache.get(key)
        if cached and time.time() - cached[0] < _TOP_CHANNELS_TTL:
            return cached[1]

    if q:
        tc_from = ("FROM messages JOIN messages_fts ON messages.id = messages_fts.rowid "
                   "WHERE messages_fts MATCH ?")
        params: list = [build_fts_query(q)]
        if since_ts:
            tc_from += " AND messages.saved_at >= ?"
            params.append(since_ts)
        if until_ts:
            tc_from += " AND messages.saved_at <= ?"
            params.append(until_ts)
        link_col = "messages.channel_link"
        title_col = "messages.channel_title"
        has_where = True
    else:
        tc_from = "FROM messages"
        params = []
        link_col = "channel_link"
        title_col = "channel_title"
        has_where = False

    if mode in ("main", "alert") and alert_aliases:
        alert_links = [f"https://max.ru/{a}".lower() for a in alert_aliases]
        ph = ",".join("?" * len(alert_links))
        op = "NOT IN" if mode == "main" else "IN"
        glue = "AND" if has_where else "WHERE"
        tc_from += f" {glue} lower({link_col}) {op} ({ph})"
        params.extend(alert_links)

    sql = (f"SELECT {title_col} AS channel_title, {link_col} AS channel_link, "
           f"       COUNT(*) AS cnt {tc_from} "
           f"GROUP BY {title_col} ORDER BY cnt DESC LIMIT 20")

    try:
        rows = db.execute(sql, params).fetchall()
    except sqlite3.OperationalError:
        rows = []
    with _top_channels_lock:
        _top_channels_cache[key] = (time.time(), rows)
    return rows


def get_top_channels_total(db, q: str, since_ts: str | None, until_ts: str | None,
                           mode: str = "all") -> int:
    """Загальна кількість унікальних каналів у вибірці (не обмежена LIMIT 20)."""
    alert_aliases = _load_alert_channels() if mode != "all" else set()
    if mode == "alert" and not alert_aliases:
        return 0

    if q:
        tc_from = ("FROM messages JOIN messages_fts ON messages.id = messages_fts.rowid "
                   "WHERE messages_fts MATCH ?")
        params: list = [build_fts_query(q)]
        if since_ts:
            tc_from += " AND messages.saved_at >= ?"; params.append(since_ts)
        if until_ts:
            tc_from += " AND messages.saved_at <= ?"; params.append(until_ts)
        link_col = "messages.channel_link"
        title_col = "messages.channel_title"
        has_where = True
    else:
        tc_from = "FROM messages"
        params = []
        link_col = "channel_link"
        title_col = "channel_title"
        has_where = False

    if mode in ("main", "alert") and alert_aliases:
        alert_links = [f"https://max.ru/{a}".lower() for a in alert_aliases]
        ph = ",".join("?" * len(alert_links))
        op = "NOT IN" if mode == "main" else "IN"
        glue = "AND" if has_where else "WHERE"
        tc_from += f" {glue} lower({link_col}) {op} ({ph})"
        params.extend(alert_links)

    try:
        return db.execute(f"SELECT COUNT(DISTINCT {title_col}) {tc_from}", params).fetchone()[0] or 0
    except sqlite3.OperationalError:
        return 0

def get_stats(db) -> dict:
    global _stats_cache
    with _stats_lock:
        if _stats_cache and time.time() - _stats_cache[0] < _STATS_CACHE_TTL:
            return _stats_cache[1]
    now = datetime.now()
    hour_ago      = (now - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
    two_hours_ago = (now - timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S")
    day_ago       = (now - timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
    two_days_ago  = (now - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")
    stats = {
        "total":     db.execute("SELECT COUNT(*) FROM messages").fetchone()[0],
        "channels":  db.execute("SELECT COUNT(DISTINCT channel_link) FROM messages").fetchone()[0],
        "last_hour": db.execute("SELECT COUNT(*) FROM messages WHERE saved_at >= ?", (hour_ago,)).fetchone()[0],
        "last_day":  db.execute("SELECT COUNT(*) FROM messages WHERE saved_at >= ?", (day_ago,)).fetchone()[0],
        # Постові дельти проти попереднього аналогічного періоду — для стрілок
        # у картках "За годину" / "За 24 год".
        "last_hour_prev": db.execute(
            "SELECT COUNT(*) FROM messages WHERE saved_at >= ? AND saved_at < ?",
            (two_hours_ago, hour_ago)).fetchone()[0],
        "last_day_prev":  db.execute(
            "SELECT COUNT(*) FROM messages WHERE saved_at >= ? AND saved_at < ?",
            (two_days_ago, day_ago)).fetchone()[0],
        # Активні унікальні канали за 24h + попередня доба — для бейджа в "Каналів".
        "active_channels_24h":      db.execute(
            "SELECT COUNT(DISTINCT channel_link) FROM messages WHERE saved_at >= ?",
            (day_ago,)).fetchone()[0],
        "active_channels_24h_prev": db.execute(
            "SELECT COUNT(DISTINCT channel_link) FROM messages WHERE saved_at >= ? AND saved_at < ?",
            (two_days_ago, day_ago)).fetchone()[0],
    }
    with _stats_lock:
        _stats_cache = (time.time(), stats)
    return stats


# ── Timeline (sparkline 30 днів) ──────────────────────────────────────────────

_TIMELINE_TTL = 60
_timeline_cache: dict[tuple, tuple[float, list]] = {}
_timeline_lock = threading.Lock()


# ── Reach (охоплення = sum(views) по днях) ────────────────────────────────────
# Збір повільний (WS до MAX), тому асинхронна модель: POST стартує task,
# GET <task_id> опитує прогрес. Кеш 15 хв на (q, channel, days).
# Глобальний семафор: одночасно лише один reach-task (login-токен MAX один).

_REACH_CACHE_TTL = 900           # 15 хв
_REACH_FULL_LIMIT = 1500         # до цієї кількості постів — повний прохід
_REACH_SAMPLE_SIZE = 800         # цільовий розмір вибірки понад FULL_LIMIT
_REACH_HARD_LIMIT = 8000         # стеля SELECT — навіть без помилки, просто обрізаємо
_REACH_ALLOWED_DAYS = (7, 30)    # 90 — занадто
_REACH_TASK_TTL = 1800           # таски прибираємо через 30 хв
_REACH_MAX_QUEUE = 10            # стеля одночасних task-ів у черзі (anti-spam)

_reach_cache: dict[tuple, tuple[float, list]] = {}
_reach_tasks: dict[str, dict] = {}
_reach_lock = threading.Lock()
_reach_queue_cond = threading.Condition(_reach_lock)
_reach_pending_ids: list[str] = []          # FIFO id-ів task-ів у стані "queued"
_reach_running: dict[str, bool] = {"busy": False}


def _reach_gc():
    """Прибрати старі таски, щоб не накопичувались у пам'яті."""
    now = time.time()
    with _reach_lock:
        stale = [tid for tid, t in _reach_tasks.items()
                 if now - t.get("ts_done", t["ts_started"]) > _REACH_TASK_TTL]
        for tid in stale:
            _reach_tasks.pop(tid, None)


def _stratified_sample(
    rows: list[dict], target: int
) -> tuple[list[dict], dict[str, tuple[int, int]]]:
    """Стратифікована вибірка: пропорційно по днях, мінімум до 5 постів на день
    (якщо стільки взагалі є). Повертає (вибірка, {date: (day_total, day_sampled)})."""
    by_day: dict[str, list[dict]] = {}
    for r in rows:
        by_day.setdefault(r["d"], []).append(r)

    total = len(rows)
    sampled_posts: list[dict] = []
    day_meta: dict[str, tuple[int, int]] = {}
    for d, day_rows in by_day.items():
        proportional = round(len(day_rows) * target / total)
        quota = max(min(5, len(day_rows)), proportional)
        quota = min(len(day_rows), quota)
        chunk = random.sample(day_rows, quota) if quota < len(day_rows) else list(day_rows)
        sampled_posts.extend(chunk)
        day_meta[d] = (len(day_rows), len(chunk))

    return sampled_posts, day_meta


def _run_reach_task(task_id: str, q: str, channel: str, days: int):
    """Worker: SELECT матчених постів → (опц. семплинг) → fetch_views → агрегація."""
    import views_fetcher

    task = _reach_tasks[task_id]
    try:
        db = get_db()
        cutoff_dt = datetime.now() - timedelta(days=days)
        cutoff = cutoff_dt.strftime("%Y-%m-%d %H:%M:%S")
        oldest_ts_ms = int(cutoff_dt.timestamp() * 1000)
        fts_q = build_fts_query(q)
        sql = ("SELECT messages.chat_id AS chat_id, messages.msg_id AS msg_id, "
               "       messages.msg_time AS msg_time, date(messages.saved_at) AS d "
               "FROM messages JOIN messages_fts ON messages.id = messages_fts.rowid "
               "WHERE messages_fts MATCH ? AND messages.saved_at >= ?")
        params: list = [fts_q, cutoff]
        if channel:
            sql += " AND messages.channel_title = ?"
            params.append(channel)
        sql += f" ORDER BY messages.saved_at DESC LIMIT {_REACH_HARD_LIMIT}"

        try:
            rows = [dict(r) for r in db.execute(sql, params).fetchall()]
        except sqlite3.OperationalError as e:
            task.update(state="error", error=f"SQL: {e}", ts_done=time.time())
            return
        finally:
            db.close()

        if not rows:
            task.update(state="done", data=[], ts_done=time.time(),
                        posts_total=0, posts_sampled=0, sampled=False)
            return

        posts_total = len(rows)
        sampled_mode = posts_total > _REACH_FULL_LIMIT
        if sampled_mode:
            sampled_posts, day_meta = _stratified_sample(rows, _REACH_SAMPLE_SIZE)
        else:
            sampled_posts = rows
            day_meta = {}

        # Прогрес з логів views_fetcher: "[views] N/M chat=..."
        progress_re = re.compile(r"\[views\]\s+(\d+)/(\d+)\s+chat=")

        def log_capture(msg):
            print(msg, flush=True)
            m = progress_re.search(str(msg))
            if m:
                task["progress"] = int(m.group(1))
                task["total"] = int(m.group(2))

        task["total"] = len({r["chat_id"] for r in sampled_posts})
        task["progress"] = 0
        task["state"] = "running"

        posts_for_views = [
            {"chat_id": r["chat_id"], "msg_id": r["msg_id"], "msg_time": r["msg_time"]}
            for r in sampled_posts
        ]
        views_map = views_fetcher.fetch_views(
            posts_for_views, log=log_capture, oldest_ts_ms=oldest_ts_ms
        )

        # Сума переглядів по днях у вибірці
        sampled_sums: dict[str, int] = {}
        for r in sampled_posts:
            v = views_map.get((r["chat_id"], str(r["msg_id"])))
            if isinstance(v, int):
                sampled_sums[r["d"]] = sampled_sums.get(r["d"], 0) + v

        today = datetime.now().date()
        result = []
        for i in range(days - 1, -1, -1):
            d = (today - timedelta(days=i)).isoformat()
            sv = sampled_sums.get(d, 0)
            if sampled_mode and d in day_meta:
                day_total, day_sampled = day_meta[d]
                views_d = round(sv * day_total / day_sampled) if day_sampled > 0 else 0
            else:
                views_d = sv
            result.append({"date": d, "views": views_d})

        cache_key = (q, channel, days)
        with _reach_lock:
            _reach_cache[cache_key] = (time.time(), result)

        task.update(state="done", data=result, ts_done=time.time(),
                    posts_total=posts_total, posts_sampled=len(sampled_posts),
                    sampled=sampled_mode)
    except Exception as e:
        task.update(state="error", error=f"exception: {e}", ts_done=time.time())
    finally:
        _reach_running["busy"] = False


# ── AI-аналітика через Claude API ─────────────────────────────────────────────
# Promt з summary.txt + усі пости за поточним фільтром (q + period). Виклик
# повільний (15-30с на Opus), тому async-task pattern як у reach. Глобальний
# семафор: один Claude-запит одночасно.

_analytics_cache:   dict[tuple, tuple[float, dict]] = {}
_analytics_tasks:   dict[str, dict] = {}
_analytics_lock     = threading.Lock()
_analytics_running: dict[str, bool] = {"busy": False}


def _analytics_gc():
    now = time.time()
    with _analytics_lock:
        stale = [tid for tid, t in _analytics_tasks.items()
                 if now - t.get("ts_done", t["ts_started"]) > ANALYTICS_TASK_TTL]
        for tid in stale:
            _analytics_tasks.pop(tid, None)


def _load_anthropic_client() -> tuple[object | None, str | None]:
    """Lazy import + ключ з файлу. Якщо є .anthropic_gateway — використовує
    base_url проксі (для обходу гео-блоку). Повертає (client, error_code)."""
    if not ANTHROPIC_KEY_FILE.exists():
        return None, "key_missing"
    key = ANTHROPIC_KEY_FILE.read_text(encoding="utf-8").strip()
    if not key:
        return None, "key_empty"
    try:
        from anthropic import Anthropic
    except ImportError:
        return None, "sdk_missing"
    kwargs: dict = {"api_key": key}
    if ANTHROPIC_GATEWAY_FILE.exists():
        gw = ANTHROPIC_GATEWAY_FILE.read_text(encoding="utf-8").strip()
        if gw:
            kwargs["base_url"] = gw
            if ANTHROPIC_GATEWAY_TOKEN_FILE.exists():
                tok = ANTHROPIC_GATEWAY_TOKEN_FILE.read_text(encoding="utf-8").strip()
                if tok:
                    kwargs["default_headers"] = {"cf-aig-authorization": f"Bearer {tok}"}
    return Anthropic(**kwargs), None


def _build_analytics_input(rows: list[sqlite3.Row]) -> tuple[str, int]:
    """Серіалізує пости у блок тексту для Claude. Дуже довгі обрізає, зупиняється
    при досягненні ANALYTICS_MAX_INPUT_CHARS. Повертає (text, used_count)."""
    parts: list[str] = []
    used = 0
    total_chars = 0
    for r in rows:
        text = (r["text"] or "").strip()
        if not text:
            continue
        if len(text) > ANALYTICS_POST_TRIM_CHARS:
            text = text[:ANALYTICS_POST_TRIM_CHARS] + "…"
        block = f"[{r['msg_time']}] {r['channel_title']}\n{text}\n---\n"
        if total_chars + len(block) > ANALYTICS_MAX_INPUT_CHARS:
            break
        parts.append(block)
        total_chars += len(block)
        used += 1
    return "".join(parts), used


def _md_to_html(md: str) -> str:
    """Мінімальний md→html для відповіді Claude (## заголовки, **bold**,
    нумеровані списки, абзаци). HTML-escape всього іншого. Без зовнішніх залежностей."""
    from html import escape
    out: list[str] = []
    para: list[str] = []

    def flush():
        if para:
            txt = " ".join(para)
            txt = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", txt)
            out.append(f"<p>{txt}</p>")
            para.clear()

    for raw in md.splitlines():
        line = escape(raw.rstrip())
        if not line.strip():
            flush()
            continue
        if line.startswith("## "):
            flush()
            out.append(f"<h3>{line[3:]}</h3>")
        elif re.match(r"^\d+\.\s", line):
            flush()
            line_html = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", line)
            out.append(f'<div class="ai-list-item">{line_html}</div>')
        else:
            para.append(line)
    flush()
    return "\n".join(out)


def _run_analytics_task(task_id: str, q: str, channel: str,
                        since_ts: str | None, until_ts: str | None):
    task = _analytics_tasks[task_id]
    try:
        db = get_db()
        try:
            rows = _query_export_rows(db, q, channel, since_ts, until_ts, sort="new")
        finally:
            db.close()

        if not rows:
            task.update(state="error", error="no_posts", ts_done=time.time())
            return

        posts_total = len(rows)
        posts_text, posts_used = _build_analytics_input(rows)
        if posts_used == 0:
            task.update(state="error", error="empty_texts", ts_done=time.time())
            return

        try:
            system_prompt = SUMMARY_PROMPT_FILE.read_text(encoding="utf-8").strip()
        except OSError as e:
            task.update(state="error", error=f"prompt_read: {e}", ts_done=time.time())
            return

        client, err = _load_anthropic_client()
        if err:
            task.update(state="error", error=err, ts_done=time.time())
            return

        period_descr = ""
        if since_ts and until_ts:
            period_descr = f"з {since_ts} по {until_ts}"
        elif since_ts:
            period_descr = f"з {since_ts} по теперішній час"
        else:
            period_descr = "за весь час"

        user_msg = (
            f"Ключове слово пошуку: «{q}»\n"
            f"Канал: {channel or '(усі канали)'}\n"
            f"Період: {period_descr}\n"
            f"Постів у вибірці: {posts_used} з {posts_total}\n\n"
            f"=== ПОСТИ ===\n{posts_text}"
        )

        task["state"] = "running"
        print(f"[analytics] start: q={q!r} channel={channel!r} posts={posts_used}/{posts_total} chars={len(posts_text)}",
              flush=True)
        t0 = time.time()
        resp = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=2_500,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        elapsed = time.time() - t0
        answer = resp.content[0].text
        print(f"[analytics] done in {elapsed:.1f}s, answer={len(answer)} chars", flush=True)

        result = {
            "markdown": answer,
            "html": _md_to_html(answer),
            "posts_total": posts_total,
            "posts_used": posts_used,
            "model": ANTHROPIC_MODEL,
            "elapsed_sec": round(elapsed, 1),
        }
        task.update(state="done", data=result, ts_done=time.time())
        cache_key = (q, channel, since_ts or "", until_ts or "")
        with _analytics_lock:
            _analytics_cache[cache_key] = (time.time(), result)
    except Exception as e:
        print(f"[analytics] error: {e}", flush=True)
        task.update(state="error", error=f"exception: {e}", ts_done=time.time())
    finally:
        _analytics_running["busy"] = False


_ALLOWED_TIMELINE_DAYS = (7, 30, 90)


def get_timeline(db, q: str, channel: str, days: int = 30) -> list[dict]:
    """
    Повертає [{date: 'YYYY-MM-DD', count: N}] за останні N днів (7/30/90).
    Часовий фільтр (period з UI) НЕ застосовуємо — графік показує контекст.
    Канал/q — застосовуємо, бо це частина фільтра пошуку.
    Кешується на 60с.
    """
    if days not in _ALLOWED_TIMELINE_DAYS:
        days = 30
    cache_key = (q, channel, days)
    with _timeline_lock:
        cached = _timeline_cache.get(cache_key)
        if cached and time.time() - cached[0] < _TIMELINE_TTL:
            return cached[1]

    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")

    if q:
        fts_q = build_fts_query(q)
        sql = ("SELECT date(messages.saved_at) AS d, COUNT(*) AS c "
               "FROM messages JOIN messages_fts ON messages.id = messages_fts.rowid "
               "WHERE messages_fts MATCH ? AND messages.saved_at >= ?")
        params: list = [fts_q, cutoff]
        if channel:
            sql += " AND messages.channel_title = ?"
            params.append(channel)
    else:
        sql = "SELECT date(saved_at) AS d, COUNT(*) AS c FROM messages WHERE saved_at >= ?"
        params = [cutoff]
        if channel:
            sql += " AND channel_title = ?"
            params.append(channel)

    sql += " GROUP BY d ORDER BY d"

    try:
        rows = db.execute(sql, params).fetchall()
    except sqlite3.OperationalError:
        rows = []

    by_date = {r[0]: r[1] for r in rows}
    today = datetime.now().date()
    result = []
    for i in range(days - 1, -1, -1):
        d = (today - timedelta(days=i)).isoformat()
        result.append({"date": d, "count": by_date.get(d, 0)})

    with _timeline_lock:
        _timeline_cache[cache_key] = (time.time(), result)
    return result


def render_sparkline_svg(timeline: list[dict], highlight_days: int = 0) -> str:
    """
    Інлайновий SVG-sparkline. highlight_days — скільки останніх барів підсвітити.
    """
    if not timeline:
        return ""
    width, height = 300, 44
    bar_w = width / len(timeline)
    max_c = max((d["count"] for d in timeline), default=1) or 1
    bars = []
    n = len(timeline)
    for i, d in enumerate(timeline):
        h = max(1, (d["count"] / max_c) * (height - 4))
        x = i * bar_w + 1
        y = height - h
        is_highlighted = highlight_days > 0 and i >= n - highlight_days
        fill = "#8b83ff" if is_highlighted else "rgba(139,131,255,0.35)"
        bars.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w-1.5:.1f}" height="{h:.1f}" '
            f'rx="1" fill="{fill}"><title>{d["date"]}: {d["count"]}</title></rect>'
        )
    return (f'<svg class="sparkline" viewBox="0 0 {width} {height}" '
            f'width="{width}" height="{height}">{"".join(bars)}</svg>')


# Додаткові російські стоп-слова поверх NLTK — ті, що Natasha POS-тегами не відсіює
# (ввідні слова, модальні, часові маркери, службові новинні конструкції,
# побутові іменники, які пробивались у топ).
_RU_EXTRA = {
    # Займенники / службові
    "это","этот","эта","эти","тот","та","те","такой","такая","такие","весь","вся",
    "все","всё","каждый","другой","сам","свой","наш","ваш","который",
    "также","тоже","уже","ещё","еще","лишь","только","даже","просто","вот","ведь",
    "может","можно","нужно","нельзя","должен","должна","должно","должны",
    "сегодня","вчера","завтра","сейчас","тогда","теперь","всегда","никогда",
    "очень","много","мало","более","менее","почти","совсем","именно","вообще",
    # Часові маркери
    "год","день","время","минута","час","неделя","месяц","утро","вечер","ночь",
    "период","сутки","момент","эпоха","эра","раз","разом",
    # Загальні новинні конструкції
    "сторона","вопрос","случай","область","район","место","часть","группа","ряд",
    "тема","новость","новости","информация","сведения","данные","детали","итог",
    "результат","причина","повод","факт","версия","мнение","комментарий",
    # Побутові іменники, що пробивались у топ
    "автомобиль","машина","авто","транспорт","грузовик","автобус","поезд","метро",
    "снег","дождь","туман","ветер","погода","температура","мороз","жара",
    "память","фото","видео","картинка","кадр","снимок","ролик","клип",
    "отель","гостиница","ресторан","кафе","магазин","рынок","супермаркет",
    "ребёнок","ребенок","дети","малыш","школьник","ученик","студент","учитель",
    "праздник","подарок","свадьба","юбилей","день_рождения",
    "телефон","смартфон","компьютер","ноутбук","экран",
    "квартира","дом","здание","комната","офис","этаж","балкон",
    "улица","дорога","шоссе","трасса","перекрёсток","перекресток","тротуар",
    "семья","родственник","друг","сосед","знакомый",
    "одежда","обувь","костюм","рубашка","платье","сумка",
    "еда","продукт","блюдо","напиток","кофе","чай","вода",
    "собака","кот","кошка","животное","птица","рыба",
    # Загальні дії-іменники (девербативи, які лишаються після фільтра POS)
    "сообщение","сообщения","заявление","заявления","выступление","выступления",
    "обращение","обращения","рассказ","разговор","беседа","интервью","репортаж",
    "слова","фраза","речь","комментарий","ответ","реакция","оценка",
}

EXTRA_STOPS = _RU_STOPS | _RU_EXTRA

_top_words_lock = threading.Lock()
_baseline_state = {"last_built": 0.0, "df": {}, "n_docs": 0}
_baseline_state_lock = threading.Lock()


def _get_baseline_snapshot(db) -> tuple[dict, int]:
    """Lazy-load baseline з БД при першому виклику, потім тримаємо в пам'яті."""
    with _baseline_state_lock:
        if _baseline_state["n_docs"] > 0:
            return _baseline_state["df"], _baseline_state["n_docs"]
    df, n_docs = nlp_mod.load_baseline(db)
    with _baseline_state_lock:
        _baseline_state["df"] = df
        _baseline_state["n_docs"] = n_docs
    return df, n_docs


_EMPTY_TOP = {c: [] for c in nlp_mod.CATEGORIES}


def _compute_top_words_blocking(period: str, channel: str = "") -> dict:
    """Реальне обчислення (NER+TF-IDF). Викликається у фоновому thread.
    Якщо channel — рахує тільки по постах цього каналу (channel_title).

    Стратегія:
      1) Спочатку читаємо TF з `message_lemmas` (агрегатний SQL — мс).
      2) Якщо кеш покриває >= LEMMA_CACHE_MIN_COVERAGE постів періоду — повертаємо.
      3) Інакше — fallback на старий on-the-fly NER (з дедуплікацією).
    Background scheduler (`lemma_cache_scheduler`) у будь-якому разі догризає
    pending-пости, тож наступні запити будуть швидші.
    """
    delta = {"day": timedelta(days=1), "week": timedelta(weeks=1), "month": timedelta(days=30)}
    since = (datetime.now() - delta.get(period, timedelta(days=1))).strftime("%Y-%m-%d %H:%M:%S")
    extra_stops = EXTRA_STOPS | load_custom_stops()
    db = get_db()
    try:
        tf_period, done_in_period, total_in_period = nlp_mod.compute_period_tf_from_cache(
            db, since, extra_stops=extra_stops, channel=channel or None
        )
        coverage = (done_in_period / total_in_period) if total_in_period else 1.0
        if total_in_period and coverage < LEMMA_CACHE_MIN_COVERAGE:
            print(f"[top-words] cache coverage {coverage:.0%} ({done_in_period}/{total_in_period}) "
                  f"< {LEMMA_CACHE_MIN_COVERAGE:.0%}; fallback на on-the-fly NER", flush=True)
            tf_period = nlp_mod.compute_period_tf(
                db, since, MAX_ROWS_SCAN, extra_stops, channel=channel or None
            )
        else:
            print(f"[top-words] from cache: coverage={coverage:.0%} "
                  f"({done_in_period}/{total_in_period} постів)", flush=True)
        baseline_df, n_baseline = _get_baseline_snapshot(db)
    finally:
        db.close()
    return nlp_mod.score_top_categorized(
        tf_period, baseline_df, n_baseline, min_tf=3, limit_per_cat=200
    )


def _start_top_words_compute(period: str, channel: str = ""):
    """Запускає фонове обчислення топу для (період, канал) — одне на пару одночасно."""
    key = (period, channel)
    with _top_words_lock:
        if key in _top_words_inflight:
            return
        _top_words_inflight.add(key)

    def _worker():
        label = f"{period}/{channel}" if channel else period
        try:
            t0 = time.time()
            by_cat = _compute_top_words_blocking(period, channel)
            with _top_words_lock:
                _top_words_cache[key] = (time.time(), by_cat)
            print(f"[top-words] {label}: готово за {time.time()-t0:.1f}s", flush=True)
        except Exception as e:
            print(f"[top-words] {label}: помилка — {e}", flush=True)
        finally:
            with _top_words_lock:
                _top_words_inflight.discard(key)

    threading.Thread(target=_worker, daemon=True).start()


def get_top_words(db, period: str, channel: str = "", limit_per_cat: int = 10):
    """
    Повертає dict {per:[...], loc:[...], org:[...], term:[...]}.
    Неблокуюча: якщо кешу немає — стартує фонове обчислення і одразу повертає
    порожній dict. UI перезапитає через API за хвилину і отримає дані.

    channel="" — глобальна статистика; інакше — тільки по конкретному каналу.
    """
    if not nlp_mod.nlp_available():
        return dict(_EMPTY_TOP)

    key = (period, channel)
    with _top_words_lock:
        cached = _top_words_cache.get(key)
        if cached and time.time() - cached[0] < TOP_WORDS_CACHE_TTL:
            return {c: cached[1][c][:limit_per_cat] for c in nlp_mod.CATEGORIES}

    _start_top_words_compute(period, channel)
    return dict(_EMPTY_TOP)


def _rebuild_baseline_once():
    """Синхронна перебудова baseline. Викликається з фонового потоку."""
    db = get_db()
    try:
        extra_stops = EXTRA_STOPS | load_custom_stops()
        n_docs, n_lemmas = nlp_mod.build_baseline(db, days_back=30, extra_stops=extra_stops)
        print(f"[baseline] rebuilt: {n_docs} docs, {n_lemmas} unique lemmas", flush=True)
        # Скинути snapshot щоб наступний запит перечитав з БД
        with _baseline_state_lock:
            _baseline_state["df"] = {}
            _baseline_state["n_docs"] = 0
            _baseline_state["last_built"] = time.time()
        # Скинути кеш топів
        with _top_words_lock:
            _top_words_cache.clear()
    except Exception as e:
        print(f"[baseline] rebuild failed: {e}", flush=True)
    finally:
        db.close()


def _read_baseline_built_at(db) -> float:
    """Повертає UNIX-час останнього оновлення baseline з BD (0.0 якщо порожнє)."""
    try:
        row = db.execute(
            f"SELECT MAX(updated_at) FROM {nlp_mod.BASELINE_TABLE}"
        ).fetchone()
    except sqlite3.OperationalError:
        return 0.0
    if not row or not row[0]:
        return 0.0
    try:
        return datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S").timestamp()
    except ValueError:
        return 0.0


def baseline_scheduler():
    """Фоновий потік: перевіряє раз на годину чи треба перебудувати baseline."""
    def _loop():
        time.sleep(30)  # дати серверу стартонути
        while True:
            try:
                db = get_db()
                _, n_docs = nlp_mod.load_baseline(db)
                built_at = _read_baseline_built_at(db) if n_docs > 0 else 0.0
                db.close()
                with _baseline_state_lock:
                    if built_at > _baseline_state["last_built"]:
                        _baseline_state["last_built"] = built_at
                age = time.time() - _baseline_state["last_built"]
                need_rebuild = (n_docs == 0) or (age > BASELINE_REBUILD_INTERVAL)
                if need_rebuild:
                    print("[baseline] starting rebuild...", flush=True)
                    _rebuild_baseline_once()
            except Exception as e:
                print(f"[baseline scheduler] {e}", flush=True)
            time.sleep(3600)  # перевіряти раз на годину
    threading.Thread(target=_loop, daemon=True).start()


def warm_top_words_cache():
    """Прогрів кешу: послідовно day → week → month після старту сервера.
    Прогріваємо тільки глобальну статистику (channel=""); per-channel — по запиту."""
    def _warm():
        time.sleep(5)
        for period in ("day", "week", "month"):
            key = (period, "")
            try:
                _start_top_words_compute(period, "")
                # Чекаємо завершення цього періоду перш ніж стартувати наступний,
                # щоб не запустити 3 важких обчислення паралельно.
                while key in _top_words_inflight:
                    time.sleep(1)
            except Exception as e:
                print(f"warm {period}: {e}", flush=True)
    threading.Thread(target=_warm, daemon=True).start()


def lemma_cache_scheduler():
    """Фоновий потік: інкрементально наповнює `message_lemmas` для нових постів.

    Стратегія пріоритету:
      - Спершу обробляємо пости періоду "за 24 години" (top-words основний UX).
      - Коли цей період повністю в кеші — догризаємо решту бази (week, month, ...).
      - Коли все оброблено — спимо LEMMA_CACHE_INTERVAL до наступного проходу.

    Workers контролюються env-змінною `MAX_PARSER_NLP_WORKERS`. У `1` (default)
    обробка sequential і безпечна по RAM. >1 запускає ProcessPoolExecutor —
    кожен worker тримає Natasha-моделі (~600 МБ); підняти ліміт у systemd.
    """
    def _loop():
        time.sleep(8)  # дати warm_top_words_cache стартонути першим
        # Перший прохід — лога-friendly статус.
        db = get_db()
        try:
            done, total = nlp_mod.lemma_cache_progress(db)
            print(f"[lemma-cache] стартую: {done}/{total} постів вже в кеші, "
                  f"workers={LEMMA_CACHE_WORKERS}", flush=True)
        finally:
            db.close()

        while True:
            try:
                extra_stops = EXTRA_STOPS | load_custom_stops()
                # Пріоритет — період останньої доби.
                since_24h = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
                db = get_db()
                try:
                    t0 = time.time()
                    n = nlp_mod.process_messages_batch(
                        db, LEMMA_CACHE_BATCH_SIZE, extra_stops,
                        since=since_24h, n_workers=LEMMA_CACHE_WORKERS
                    )
                    if n == 0:
                        # 24h в кеші — добираємо решту (без `since`).
                        n = nlp_mod.process_messages_batch(
                            db, LEMMA_CACHE_BATCH_SIZE, extra_stops,
                            since=None, n_workers=LEMMA_CACHE_WORKERS
                        )
                    if n > 0:
                        elapsed = time.time() - t0
                        rate = n / elapsed if elapsed > 0 else 0.0
                        done, total = nlp_mod.lemma_cache_progress(db)
                        print(f"[lemma-cache] +{n} постів за {elapsed:.1f}с "
                              f"({rate:.1f}/с) → {done}/{total}", flush=True)
                finally:
                    db.close()

                # Якщо щось обробили — швидко далі; інакше пауза.
                time.sleep(LEMMA_CACHE_BUSY_INTERVAL if n > 0 else LEMMA_CACHE_INTERVAL)
            except Exception as e:
                print(f"[lemma-cache] error: {e}", flush=True)
                time.sleep(LEMMA_CACHE_INTERVAL)

    threading.Thread(target=_loop, daemon=True).start()


@app.route("/api/add-stop-word", methods=["POST"])
def api_add_stop_word():
    word = (request.json or {}).get("word", "").strip().lower()
    if not word or len(word) < 2:
        return jsonify({"ok": False, "error": "empty"}), 400
    existing = load_custom_stops()
    if word not in existing:
        save_custom_stop(word)
    # Видаляємо лему з інкрементального кеша — інакше вона ще покажеться у топі
    # до повного перебору. Робимо в окремому коннекті, не блокуючи відповідь надовго.
    db = get_db()
    try:
        deleted = nlp_mod.purge_lemma_from_cache(db, word)
    finally:
        db.close()
    if deleted:
        print(f"[stop-word] '{word}': видалено {deleted} рядків з {nlp_mod.LEMMA_CACHE_TABLE}",
              flush=True)
    with _top_words_lock:
        _top_words_cache.clear()
    return jsonify({"ok": True, "word": word})


@app.route("/api/timeline")
def api_timeline():
    q       = request.args.get("q", "").strip()
    channel = request.args.get("channel", "").strip()
    try:
        days = int(request.args.get("days", 30))
    except ValueError:
        days = 30
    if days not in _ALLOWED_TIMELINE_DAYS:
        days = 30
    db = get_db()
    try:
        data = get_timeline(db, q, channel, days)
    finally:
        db.close()
    return jsonify(data)


def _reach_queue_position_locked(task_id: str) -> int | None:
    """Кількість task-ів попереду у черзі (без поточного). 0 = наступний.
    `None` якщо task не серед очікуючих. Викликати лише під _reach_lock."""
    ahead = 1 if _reach_running["busy"] else 0
    for tid in _reach_pending_ids:
        if tid == task_id:
            return ahead
        ahead += 1
    return None


def reach_dispatcher():
    """Фоновий serial-worker: бере наступний task з черги і виконує його.

    Серіалізація потрібна бо MAX блокує паралельні WS-сесії з тим самим
    LOGIN-токеном (див. CLAUDE.md). Раніше API повертало 429 на другий запит;
    тепер усі запити стають у чергу і обслуговуються по черзі — без помилки.
    """
    def _loop():
        while True:
            with _reach_queue_cond:
                while not _reach_pending_ids:
                    _reach_queue_cond.wait()
                task_id = _reach_pending_ids.pop(0)
                task = _reach_tasks.get(task_id)
                if task is None:
                    continue
                _reach_running["busy"] = True
                task["state"] = "pending"
            params = task["params"]
            # _run_reach_task сам ставить busy=False у finally
            try:
                _run_reach_task(task_id, params["q"], params["channel"], params["days"])
            except Exception as e:
                # Захист на випадок неочікуваної помилки поза try/finally воркера —
                # busy має бути скинутий, інакше черга залипне.
                print(f"[reach] dispatcher exception: {e}", flush=True)
                with _reach_lock:
                    _reach_running["busy"] = False
                    if task is not None:
                        task["state"] = "error"
                        task["error"] = f"dispatcher: {e}"
                        task["ts_done"] = time.time()
    threading.Thread(target=_loop, daemon=True).start()


@app.route("/api/timeline-reach", methods=["POST"])
def api_timeline_reach_start():
    q       = (request.json or {}).get("q", "").strip() if request.is_json else request.form.get("q", "").strip()
    channel = (request.json or {}).get("channel", "").strip() if request.is_json else request.form.get("channel", "").strip()
    try:
        days = int((request.json or {}).get("days", 30)) if request.is_json else int(request.form.get("days", 30))
    except (ValueError, TypeError):
        days = 30

    if not q:
        return jsonify({"error": "q_required"}), 400
    if days not in _REACH_ALLOWED_DAYS:
        days = 30

    cache_key = (q, channel, days)
    with _reach_lock:
        cached = _reach_cache.get(cache_key)
        if cached and time.time() - cached[0] < _REACH_CACHE_TTL:
            return jsonify({
                "task_id": "cached",
                "state": "done",
                "data": cached[1],
                "cached": True,
            })

    _reach_gc()

    with _reach_queue_cond:
        # Anti-spam: не пускаємо більше N task-ів одночасно у черзі.
        if len(_reach_pending_ids) >= _REACH_MAX_QUEUE:
            return jsonify({"error": "queue_full",
                            "queue_size": len(_reach_pending_ids)}), 429

        task_id = uuid.uuid4().hex[:12]
        _reach_tasks[task_id] = {
            "state": "queued",
            "progress": 0,
            "total": 0,
            "data": None,
            "error": None,
            "ts_started": time.time(),
            "params": {"q": q, "channel": channel, "days": days},
        }
        _reach_pending_ids.append(task_id)
        position = _reach_queue_position_locked(task_id)
        _reach_queue_cond.notify()

    return jsonify({
        "task_id": task_id,
        "state": "queued",
        "queue_position": position,
    })


@app.route("/api/timeline-reach/<task_id>")
def api_timeline_reach_status(task_id: str):
    with _reach_lock:
        task = _reach_tasks.get(task_id)
        position = _reach_queue_position_locked(task_id) if task else None
    if not task:
        return jsonify({"error": "unknown_task"}), 404
    out = {
        "state":    task["state"],
        "progress": task.get("progress", 0),
        "total":    task.get("total", 0),
    }
    if task["state"] == "queued":
        out["queue_position"] = position if position is not None else 0
    if task["state"] == "done":
        out["data"] = task["data"]
        out["posts_total"] = task.get("posts_total", 0)
        out["posts_sampled"] = task.get("posts_sampled", 0)
        out["sampled"] = task.get("sampled", False)
    elif task["state"] == "error":
        out["error"] = task.get("error", "unknown")
    return jsonify(out)


_ANALYTICS_ALLOWED_PERIODS = ("24h", "7d")


@app.route("/api/analytics", methods=["POST"])
def api_analytics_start():
    data = request.json or {}
    q       = (data.get("q") or "").strip()
    if not q:
        return jsonify({"error": "q_required"}), 400
    channel = (data.get("channel") or "").strip()
    period_in = (data.get("period") or DEFAULT_PERIOD).strip()
    if period_in not in _ANALYTICS_ALLOWED_PERIODS:
        return jsonify({"error": "period_not_allowed"}), 400
    from_date = (data.get("from_date") or "").strip()
    to_date   = (data.get("to_date") or "").strip()
    since_ts, until_ts, _ = parse_time_filter(period_in, from_date, to_date)

    cache_key = (q, channel, since_ts or "", until_ts or "")
    with _analytics_lock:
        cached = _analytics_cache.get(cache_key)
        if cached and time.time() - cached[0] < ANALYTICS_CACHE_TTL:
            return jsonify({
                "task_id": "cached", "state": "done",
                "data": cached[1], "cached": True,
            })

    _analytics_gc()

    with _analytics_lock:
        if _analytics_running["busy"]:
            return jsonify({"error": "busy", "retry_after": 5}), 429
        _analytics_running["busy"] = True
        task_id = uuid.uuid4().hex[:12]
        _analytics_tasks[task_id] = {
            "state": "pending",
            "ts_started": time.time(),
            "params": {"q": q, "channel": channel,
                       "since_ts": since_ts, "until_ts": until_ts},
        }

    threading.Thread(
        target=_run_analytics_task,
        args=(task_id, q, channel, since_ts, until_ts), daemon=True
    ).start()

    return jsonify({"task_id": task_id, "state": "pending"})


@app.route("/api/analytics/<task_id>")
def api_analytics_status(task_id: str):
    with _analytics_lock:
        task = _analytics_tasks.get(task_id)
    if not task:
        return jsonify({"error": "unknown_task"}), 404
    out = {"state": task["state"]}
    if task["state"] == "done":
        out["data"] = task["data"]
    elif task["state"] == "error":
        out["error"] = task.get("error", "unknown")
    return jsonify(out)


@app.route("/api/top-words")
def api_top_words():
    period = request.args.get("period", "day")
    if period not in ("day", "week", "month"):
        period = "day"
    channel = request.args.get("channel", "").strip()
    db = get_db()
    words = get_top_words(db, period, channel)
    db.close()
    return jsonify(words)


# ── Експорт у XLSX ────────────────────────────────────────────────────────────

_EXPORT_HEADERS = ["Канал", "Посилання", "Дата", "Кількість переглядів", "Текст"]
_EXPORT_COL_WIDTHS = [22.875, 35.875, 14.875, 20.625, 103.375]


def _query_export_rows(db, q: str, channel: str, since_ts: str | None,
                       until_ts: str | None, sort: str) -> list[sqlite3.Row]:
    """Повертає всі рядки за поточними фільтрами (без LIMIT — повна вибірка)."""
    if q:
        fts_q = build_fts_query(q)
        sql = ("SELECT messages.* FROM messages "
               "JOIN messages_fts ON messages.id = messages_fts.rowid "
               "WHERE messages_fts MATCH ?")
        params: list = [fts_q]
        if channel:
            sql += " AND messages.channel_title = ?"
            params.append(channel)
        if since_ts:
            sql += " AND messages.saved_at >= ?"
            params.append(since_ts)
        if until_ts:
            sql += " AND messages.saved_at <= ?"
            params.append(until_ts)
        sql += " ORDER BY messages.id ASC" if sort == "old" else " ORDER BY messages.id DESC"
    else:
        where, params = [], []
        if channel:
            where.append("channel_title = ?"); params.append(channel)
        if since_ts:
            where.append("saved_at >= ?"); params.append(since_ts)
        if until_ts:
            where.append("saved_at <= ?"); params.append(until_ts)
        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        order = "ORDER BY id ASC" if sort == "old" else "ORDER BY id DESC"
        sql = f"SELECT * FROM messages {where_sql} {order}"

    try:
        return db.execute(sql, params).fetchall()
    except sqlite3.OperationalError:
        return []


def _format_post_date(msg_time: str) -> str:
    """msg_time → 'DD.MM.YYYY' як у прикладі формату."""
    if not msg_time:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m %H:%M"):
        try:
            dt = datetime.strptime(msg_time, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=datetime.now().year)
            return dt.strftime("%d.%m.%Y")
        except ValueError:
            continue
    return msg_time


def _build_xlsx(q: str, period_label: str, since_ts: str | None, until_ts: str | None,
                rows: list[sqlite3.Row], views_map: dict[tuple[int, str], int | None]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Повідомлення"

    # Шапка (як у прикладі формату)
    period_str = ""
    if since_ts and until_ts:
        period_str = f"{since_ts[:10].replace('-','.')[8:10]}.{since_ts[5:7]}.{since_ts[:4]} — {until_ts[8:10]}.{until_ts[5:7]}.{until_ts[:4]}"
    if since_ts and not until_ts:
        d_from = datetime.strptime(since_ts[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
        d_to = datetime.now().strftime("%d.%m.%Y")
        period_str = f"{d_from} — {d_to}"
    if not since_ts:
        period_str = period_label or "за весь час"

    ws["A1"] = f'Запит: "{q}"' if q else "Запит: (без фільтра по тексту)"
    ws["A2"] = f"Результатів: {len(rows)}"
    ws["A3"] = f"Період: {period_str}"
    # A4 — порожній (як у прикладі)
    for i, h in enumerate(_EXPORT_HEADERS, 1):
        ws.cell(row=5, column=i, value=h)

    # Дані з 6-го рядка
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for ri, r in enumerate(rows, start=6):
        ws.cell(row=ri, column=1, value=r["channel_title"])
        link = r["post_link"] or r["channel_link"]
        c_link = ws.cell(row=ri, column=2, value=link)
        if link:
            c_link.hyperlink = link
        ws.cell(row=ri, column=3, value=_format_post_date(r["msg_time"]))
        v = views_map.get((r["chat_id"], str(r["msg_id"])))
        ws.cell(row=ri, column=4, value=v if v is not None else None)
        ws.cell(row=ri, column=5, value=r["text"]).alignment = wrap_align

    # Ширина колонок як у прикладі
    for i, w in enumerate(_EXPORT_COL_WIDTHS, 1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.route("/api/export-xlsx")
def api_export_xlsx():
    q       = request.args.get("q", "").strip()
    channel = request.args.get("channel", "").strip()
    period_in = request.args.get("period", DEFAULT_PERIOD).strip()
    if period_in not in ("1h","24h","7d","30d","all","custom"):
        period_in = DEFAULT_PERIOD
    from_date = request.args.get("from_date", "").strip()
    to_date   = request.args.get("to_date", "").strip()
    since_ts, until_ts, period = parse_time_filter(period_in, from_date, to_date)
    sort = request.args.get("sort", "new")
    if sort not in ("new", "old"):
        sort = "new"

    db = get_db()
    try:
        rows = _query_export_rows(db, q, channel, since_ts, until_ts, sort)
    finally:
        db.close()

    print(f"[export] q={q!r} channel={channel!r} period={period} rows={len(rows)}", flush=True)

    posts_for_views = [
        {"chat_id": r["chat_id"], "msg_id": r["msg_id"], "msg_time": r["msg_time"]}
        for r in rows
    ]
    views_map: dict[tuple[int, str], int | None] = {}
    if posts_for_views:
        try:
            import views_fetcher
            views_map = views_fetcher.fetch_views(
                posts_for_views, log=lambda m: print(m, flush=True))
        except Exception as e:
            print(f"[export] views fetch failed: {e}", flush=True)
            # не падаємо — віддаємо файл з порожньою колонкою переглядів

    xlsx_bytes = _build_xlsx(q, PERIOD_LABELS.get(period, ""), since_ts, until_ts, rows, views_map)

    from urllib.parse import quote
    fname_q = re.sub(r'[^\w\-_. ]', '_', q) if q else "all"
    fname = f"max_export_{fname_q}_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    fname_ascii = re.sub(r'[^\x20-\x7e]', '_', fname) or "max_export.xlsx"
    return Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="{fname_ascii}"; filename*=UTF-8\'\'{quote(fname)}'},
    )


@app.route("/")
def index():
    q            = request.args.get("q", "").strip()
    channel      = request.args.get("channel", "").strip()
    page         = max(1, int(request.args.get("page", 1)))

    period_in = request.args.get("period", DEFAULT_PERIOD).strip()
    if period_in not in ("1h","24h","7d","30d","all","custom"):
        period_in = DEFAULT_PERIOD
    from_date = request.args.get("from_date", "").strip()
    to_date   = request.args.get("to_date", "").strip()
    since_ts, until_ts, period = parse_time_filter(period_in, from_date, to_date)

    # Період для топ-слів синхронізується з основним фільтром пошуку
    _PERIOD_TO_WORDS = {"1h": "day", "24h": "day", "7d": "week",
                        "30d": "month", "all": "month", "custom": "month"}
    words_period = _PERIOD_TO_WORDS.get(period, "day")
    _WORDS_PERIOD_LABELS = {"day": "за 24 години", "week": "за 7 днів", "month": "за 30 днів"}
    words_period_label = _WORDS_PERIOD_LABELS.get(words_period, "")

    sort = request.args.get("sort", "new")
    if sort not in ("new", "old"):
        sort = "new"
    order_sql = "ORDER BY messages.id ASC" if sort == "old" else "ORDER BY messages.id DESC"

    per_page = 20

    db = get_db()

    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    stats = get_stats(db)

    # FTS5 пошук: якщо є q — через MATCH з JOIN, інакше чистий messages.
    if q:
        fts_q = build_fts_query(q)
        base_from = ("FROM messages JOIN messages_fts ON messages.id = messages_fts.rowid "
                     "WHERE messages_fts MATCH ?")
        params = [fts_q]
        if channel:
            base_from += " AND messages.channel_title = ?"
            params.append(channel)
        if since_ts:
            base_from += " AND messages.saved_at >= ?"
            params.append(since_ts)
        if until_ts:
            base_from += " AND messages.saved_at <= ?"
            params.append(until_ts)

        try:
            total_count = db.execute(f"SELECT COUNT(*) {base_from}", params).fetchone()[0]
        except sqlite3.OperationalError:
            # Невалідний FTS syntax від користувача — показуємо 0 результатів
            total_count = 0

        total_pages = max(1, (total_count + per_page - 1) // per_page)
        page = min(page, total_pages)
        offset = (page - 1) * per_page

        if total_count > 0:
            posts = db.execute(
                "SELECT messages.*, "
                "  highlight(messages_fts, 0, '<mark class=\"highlight\">', '</mark>') AS text_hl "
                f"{base_from} {order_sql} LIMIT ? OFFSET ?",
                params + [per_page, offset]
            ).fetchall()
        else:
            posts = []
    else:
        where, params = [], []
        if channel:
            where.append("channel_title = ?")
            params.append(channel)
        if since_ts:
            where.append("saved_at >= ?")
            params.append(since_ts)
        if until_ts:
            where.append("saved_at <= ?")
            params.append(until_ts)
        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        order_simple = "ORDER BY id ASC" if sort == "old" else "ORDER BY id DESC"

        total_count = db.execute(f"SELECT COUNT(*) FROM messages {where_sql}", params).fetchone()[0]
        total_pages = max(1, (total_count + per_page - 1) // per_page)
        page = min(page, total_pages)
        offset = (page - 1) * per_page

        posts = db.execute(
            f"SELECT * FROM messages {where_sql} {order_simple} LIMIT ? OFFSET ?",
            params + [per_page, offset]
        ).fetchall()

    # Список каналів для фільтру (TTL 300с — рідко змінюється)
    channel_list = get_channel_list(db)

    # Топ каналів: при пошуку — один список серед матчених; без q — два списки
    # (основний потік + alert-канали з channels/alert_channels.txt). TTL 60с.
    if q:
        top_channels = get_top_channels(db, q, since_ts, until_ts, mode="all")
        top_channels_total = get_top_channels_total(db, q, since_ts, until_ts, mode="all")
        top_channels_alert = None
        top_channels_alert_total = 0
    else:
        top_channels = get_top_channels(db, q, since_ts, until_ts, mode="main")
        top_channels_total = get_top_channels_total(db, q, since_ts, until_ts, mode="main")
        top_channels_alert = get_top_channels(db, q, since_ts, until_ts, mode="alert")
        top_channels_alert_total = get_top_channels_total(db, q, since_ts, until_ts, mode="alert")

    # Топ слів (з урахуванням обраного каналу)
    top_words = get_top_words(db, words_period, channel)

    # Timeline (sparkline): останні 30 днів
    timeline = get_timeline(db, q, channel)

    db.close()

    # Якщо q було — text_hl уже заповнено FTS5 highlight() у SELECT.
    # Без q — просто дублюємо text у text_hl (шаблон завжди рендерить text_hl).
    posts = [dict(p) for p in posts]
    for p in posts:
        if "text_hl" not in p or p["text_hl"] is None:
            p["text_hl"] = p["text"]

    # Sparkline highlight: скільки останніх барів підсвітити
    highlight_days = {"24h": 1, "7d": 7, "30d": 30}.get(period, 0)
    sparkline_svg = render_sparkline_svg(timeline, highlight_days=highlight_days)
    period_label = PERIOD_LABELS.get(period, "")
    period_dates = _format_period_dates(period, since_ts, until_ts)

    return render_template_string(TEMPLATE,
        stats=stats, posts=posts,
        q=q, channel=channel, page=page,
        total_count=total_count, total_pages=total_pages,
        channel_list=channel_list, top_channels=top_channels,
        top_channels_total=top_channels_total,
        top_channels_alert=top_channels_alert,
        top_channels_alert_total=top_channels_alert_total,
        top_words=top_words, words_period=words_period,
        nlp_ready=nlp_mod.nlp_available(),
        period=period, from_date=from_date, to_date=to_date, sort=sort,
        period_label=period_label, period_dates=period_dates,
        words_period_label=words_period_label,
        sparkline_svg=sparkline_svg,
        now=now_ts,
        current_user=session.get("user"),
    )


if __name__ == "__main__":
    _init_db = get_db()
    init_fts(_init_db)                         # FTS5 virtual table + triggers + bootstrap
    nlp_mod.init_baseline_schema(_init_db)     # baseline tables
    nlp_mod.init_lemma_cache_schema(_init_db)  # message_lemmas + done-таблиця
    _init_db.close()

    baseline_scheduler()
    warm_top_words_cache()
    lemma_cache_scheduler()
    reach_dispatcher()
    app.run(host="0.0.0.0", port=8080, debug=False, threaded=True)
